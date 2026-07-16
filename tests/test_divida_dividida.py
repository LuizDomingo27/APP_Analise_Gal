# -*- coding: utf-8 -*-
"""
Testes da camada de cobrança dividida (src/data/divida_dividida.py):
  - split_records: divisão proporcional pura (valor/minutos/quantidade);
  - save_split_charge: gravação atômica das duas metades com o mesmo código;
  - load_dividas_divididas + geração do xlsx executivo.
"""

import io
from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import text

import src.data.database as db
import src.data.divida_dividida as dd
from src.config.settings import COLS


# ── Fixture: banco isolado em SQLite temporário ───────────────────────────────

@pytest.fixture
def temp_db(tmp_path, monkeypatch):
    """Isola os módulos de banco/divisão num SQLite temporário para o teste."""
    db_path = tmp_path / "test_analise_gal.db"
    sqlite_url = f"sqlite:///{db_path}"
    monkeypatch.setenv("DATABASE_URL", sqlite_url)
    monkeypatch.setattr(db, "_database_url", lambda: sqlite_url)
    db.get_engine.clear()
    db.create_tables.clear()
    dd.create_tables()
    yield db_path
    db.get_engine.clear()
    db.create_tables.clear()


def _sample_records() -> pd.DataFrame:
    """Registros crus de uma cobrança (antes da divisão)."""
    return pd.DataFrame(
        {
            COLS["order"]:     ["100", "200"],
            COLS["date"]:      ["01/06/2026", "02/06/2026"],
            COLS["supplier"]:  ["Fornecedor A", "Fornecedor A"],
            COLS["quantity"]:  [10, 5],
            COLS["defect"]:    ["PONTO ESTOURADO", "TROCAR"],
            COLS["real_cut"]:  [10, 5],
            COLS["minutes"]:   [100.0, 50.0],
            COLS["value_brl"]: [2000.0, 1000.0],
        }
    )


# ── split_records (função pura) ───────────────────────────────────────────────

def test_split_records_50_50_divides_value_minutes():
    forn, emp = dd.split_records(_sample_records(), 0.5)

    assert list(emp[COLS["value_brl"]]) == [1000.0, 500.0]
    assert list(forn[COLS["value_brl"]]) == [1000.0, 500.0]
    assert list(emp[COLS["minutes"]]) == [50.0, 25.0]
    assert list(forn[COLS["minutes"]]) == [50.0, 25.0]


def test_split_records_value_halves_sum_to_original():
    df = _sample_records()
    forn, emp = dd.split_records(df, 0.5)
    soma = pd.to_numeric(forn[COLS["value_brl"]]) + pd.to_numeric(emp[COLS["value_brl"]])
    assert list(soma) == list(pd.to_numeric(df[COLS["value_brl"]]))


def test_split_records_custom_percentage():
    forn, emp = dd.split_records(_sample_records(), 0.4)  # empresa absorve 40%
    assert list(emp[COLS["value_brl"]]) == [800.0, 400.0]
    assert list(forn[COLS["value_brl"]]) == [1200.0, 600.0]


def test_split_records_quantity_is_rounded_and_reconciles():
    # 5 peças * 0.5 = 2.5 -> empresa arredonda para 2, fornecedor recebe 3.
    forn, emp = dd.split_records(_sample_records(), 0.5)
    assert list(emp[COLS["quantity"]]) == [5, 2]
    assert list(forn[COLS["quantity"]]) == [5, 3]
    # peças da empresa + fornecedor = total original (10 e 5)
    soma = emp[COLS["quantity"]].astype(int) + forn[COLS["quantity"]].astype(int)
    assert list(soma) == [10, 5]


def test_split_records_preserves_non_numeric_columns():
    forn, emp = dd.split_records(_sample_records(), 0.5)
    for out in (forn, emp):
        assert list(out[COLS["order"]]) == ["100", "200"]
        assert list(out[COLS["defect"]]) == ["PONTO ESTOURADO", "TROCAR"]
        assert list(out[COLS["date"]]) == ["01/06/2026", "02/06/2026"]


@pytest.mark.parametrize("perc", [-0.1, 1.1, 2.0])
def test_split_records_rejects_out_of_range(perc):
    with pytest.raises(ValueError):
        dd.split_records(_sample_records(), perc)


# ── save_split_charge (gravação atômica das duas metades) ─────────────────────

def test_save_split_charge_writes_both_tables_with_same_code(temp_db):
    forn, emp = dd.split_records(_sample_records(), 0.5)
    cod = dd.save_split_charge(
        df_fornecedor=forn,
        df_empresa=emp,
        cnpj="12.345.678/0001-90",
        data_cobranca=date(2026, 6, 1),
        data_vencimento=date(2026, 6, 21),
    )
    assert cod.startswith("PAG-")

    with db.get_connection() as conn:
        df_hist = pd.read_sql(
            text('SELECT * FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :c'),
            conn, params={"c": cod},
        )
        df_div = pd.read_sql(
            text('SELECT * FROM tb_divida_dividida WHERE "COD_LANCAMENTO" = :c'),
            conn, params={"c": cod},
        )

    # Metade do fornecedor no histórico (status Pendente), metade da empresa na
    # tb_divida_dividida (status Dividida), ambas com o mesmo código.
    assert len(df_hist) == 2
    assert len(df_div) == 2
    assert set(df_hist["STATUS_COBRANCA"]) == {"Pendente"}
    assert set(df_div["STATUS_COBRANCA"]) == {dd.STATUS_DIVIDIDA}
    assert df_hist["VALOR DO PROCESSO BRL"].sum() == pytest.approx(1500.0)
    assert df_div["VALOR DO PROCESSO BRL"].sum() == pytest.approx(1500.0)


def test_save_split_charge_uses_provided_code(temp_db):
    forn, emp = dd.split_records(_sample_records(), 0.5)
    cod = dd.save_split_charge(
        df_fornecedor=forn, df_empresa=emp, cnpj="12.345.678/0001-90",
        data_cobranca=date(2026, 6, 1), data_vencimento=date(2026, 6, 21),
        cod_lancamento="PAG-FIXED01",
    )
    assert cod == "PAG-FIXED01"


# ── load_dividas_divididas + xlsx ─────────────────────────────────────────────

def test_load_returns_none_when_empty(temp_db):
    assert dd.load_dividas_divididas() is None


def test_load_returns_rows_after_save(temp_db):
    forn, emp = dd.split_records(_sample_records(), 0.5)
    dd.save_split_charge(
        df_fornecedor=forn, df_empresa=emp, cnpj="12.345.678/0001-90",
        data_cobranca=date(2026, 6, 1), data_vencimento=date(2026, 6, 21),
    )
    dd.load_dividas_divididas.clear()

    df = dd.load_dividas_divididas()
    assert df is not None
    assert len(df) == 2
    assert set(df["STATUS_COBRANCA"]) == {dd.STATUS_DIVIDIDA}
    assert set(df["DATA_PAGAMENTO"]) == {""}


def test_generate_xlsx_none_on_empty(temp_db):
    assert dd.generate_divida_dividida_xlsx_bytes() is None


def test_generate_xlsx_total_matches_company_share(temp_db):
    forn, emp = dd.split_records(_sample_records(), 0.5)
    dd.save_split_charge(
        df_fornecedor=forn, df_empresa=emp, cnpj="12.345.678/0001-90",
        data_cobranca=date(2026, 6, 1), data_vencimento=date(2026, 6, 21),
    )

    xlsx_bytes = dd.generate_divida_dividida_xlsx_bytes()
    assert xlsx_bytes is not None

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    # Linha de total logo após os 2 registros de dados.
    total_row = dd._HEADER_OFFSET + 2 + 1
    value_col_idx = ws.max_column
    assert ws.cell(row=total_row, column=value_col_idx).value == pytest.approx(1500.0)
