# -*- coding: utf-8 -*-
"""
Testes da camada de devoluções (src/data/devolucao_history.py): leitura da
tabela devolucoes e exportação do relatório executivo em xlsx.
"""

import io

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

import src.data.database as db
import src.data.devolucao_history as dh


_INSERT_SQL = (
    "INSERT INTO devolucoes "
    '(COD_LANCAMENTO, DATA_COBRANCA, DATA_VENCIMENTO, DATA_PAGAMENTO, '
    'CNPJ_FORNECEDOR, STATUS_COBRANCA, "ORDEM MESTRE", '
    '"DATA DE PRODUÇÃO ACABAMENTO", "FORNECEDOR", "QUANTIDADE", '
    '"REMONTE", "REAL CORTADO", "MINUTOS GERADOS", "VALOR DO PROCESSO BRL") '
    "VALUES (:cod, :data_cobranca, :data_vencimento, :data_pagamento, :cnpj, "
    ":status, :om, :data_producao, :fornecedor, :qtd, :remonte, :real_cortado, "
    ":minutos, :valor)"
)


@pytest.fixture
def temp_db(tmp_path, monkeypatch):
    """Isola os módulos de banco/devolução num SQLite temporário para o teste."""
    db_path = tmp_path / "test_analise_gal.db"
    sqlite_url = f"sqlite:///{db_path}"
    monkeypatch.setenv("DATABASE_URL", sqlite_url)
    monkeypatch.setattr(db, "_database_url", lambda: sqlite_url)
    db.get_engine.clear()
    db.create_tables.clear()
    dh.create_tables()
    yield db_path
    db.get_engine.clear()
    db.create_tables.clear()


def _insert_devolucao(cod, cnpj, supplier, om, qtd, remonte, real_cortado, minutos, valor):
    with db.get_connection() as conn:
        conn.execute(
            text(_INSERT_SQL),
            {
                "cod": cod, "data_cobranca": "01/06/2026", "data_vencimento": "21/06/2026",
                "data_pagamento": "", "cnpj": cnpj, "status": "Devolução", "om": om,
                "data_producao": "01/06/2026", "fornecedor": supplier, "qtd": qtd,
                "remonte": remonte, "real_cortado": real_cortado, "minutos": minutos,
                "valor": valor,
            },
        )
        conn.commit()


# ── load_devolucoes ──────────────────────────────────────────────────────────────

def test_load_devolucoes_returns_none_when_empty(temp_db):
    assert dh.load_devolucoes() is None


def test_load_devolucoes_returns_rows(temp_db):
    _insert_devolucao("PAG-DEV0001", "11.111.111/0001-11", "Fornecedor A",
                       "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)
    dh.load_devolucoes.clear()

    df = dh.load_devolucoes()
    assert df is not None
    assert len(df) == 1
    assert df.iloc[0]["COD_LANCAMENTO"] == "PAG-DEV0001"
    assert df.iloc[0]["STATUS_COBRANCA"] == "Devolução"
    assert df.iloc[0]["DATA_PAGAMENTO"] == ""


# ── generate_devolucoes_xlsx_bytes ─────────────────────────────────────────────

def test_generate_devolucoes_xlsx_bytes_returns_none_on_empty_database(temp_db):
    assert dh.generate_devolucoes_xlsx_bytes() is None


def test_generate_devolucoes_xlsx_bytes_writes_all_rows(temp_db):
    _insert_devolucao("PAG-DEV0001", "11.111.111/0001-11", "Fornecedor A",
                       "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)
    _insert_devolucao("PAG-DEV0002", "22.222.222/0001-22", "Fornecedor B",
                       "200", 3, "TROCAR", 3, 1.0, 30.0)

    xlsx_bytes = dh.generate_devolucoes_xlsx_bytes()
    assert xlsx_bytes is not None

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    all_codes_in_sheet = [
        ws.cell(row=r, column=1).value
        for r in range(dh._HEADER_OFFSET + 1, dh._HEADER_OFFSET + 3)
    ]
    assert "PAG-DEV0001" in all_codes_in_sheet
    assert "PAG-DEV0002" in all_codes_in_sheet


def test_generate_devolucoes_xlsx_bytes_totals_match_sum_of_values(temp_db):
    _insert_devolucao("PAG-DEV0001", "11.111.111/0001-11", "Fornecedor A",
                       "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)
    _insert_devolucao("PAG-DEV0002", "22.222.222/0001-22", "Fornecedor B",
                       "200", 3, "TROCAR", 3, 1.0, 30.0)

    xlsx_bytes = dh.generate_devolucoes_xlsx_bytes()
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Linha de total fica logo após os dados (2 registros).
    total_row = dh._HEADER_OFFSET + 2 + 1
    value_col_idx = ws.max_column
    assert ws.cell(row=total_row, column=value_col_idx).value == pytest.approx(80.0)
