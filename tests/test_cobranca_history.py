# -*- coding: utf-8 -*-
"""
Testes da camada de histórico de cobranças (src/data/cobranca_history.py):
badges de status/situação, agrupamento por lançamento e exportação de
xlsx de um único código (usados pela tela de Histórico de Cobranças).
"""

import io
from datetime import date, timedelta

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import text

import src.data.cobranca_history as ch
import src.data.database as db
from src.config.settings import COLS


# ── status_badge_html ──────────────────────────────────────────────────────────

def test_status_badge_html_pago():
    assert "Pago" in ch.status_badge_html("Pago")


def test_status_badge_html_devolucao():
    assert "Devolução" in ch.status_badge_html("Devolução")


def test_status_badge_html_pendente():
    assert "Pendente" in ch.status_badge_html("Pendente")


def test_status_badge_html_defaults_to_pendente_for_unknown_value():
    assert "Pendente" in ch.status_badge_html("valor-desconhecido")


# ── situacao_badge_html ─────────────────────────────────────────────────────────

def test_situacao_badge_html_pago_no_prazo():
    html = ch.situacao_badge_html("Pago", "10/06/2026", "05/06/2026")
    assert "no prazo" in html


def test_situacao_badge_html_pago_com_atraso():
    html = ch.situacao_badge_html("Pago", "10/06/2026", "15/06/2026")
    assert "atraso" in html


def test_situacao_badge_html_pago_sem_data_pagamento():
    html = ch.situacao_badge_html("Pago", "10/06/2026", "")
    assert "Informe a data do pagamento" in html


def test_situacao_badge_html_pendente_vencido():
    vencida = (date.today() - timedelta(days=3)).strftime("%d/%m/%Y")
    html = ch.situacao_badge_html("Pendente", vencida, "")
    assert "Vencido há 3d" in html


def test_situacao_badge_html_pendente_vence_hoje():
    hoje = date.today().strftime("%d/%m/%Y")
    html = ch.situacao_badge_html("Pendente", hoje, "")
    assert "Vence hoje" in html


def test_situacao_badge_html_pendente_futuro():
    futura = (date.today() + timedelta(days=5)).strftime("%d/%m/%Y")
    html = ch.situacao_badge_html("Pendente", futura, "")
    assert "5 dia(s)" in html


# ── group_charges ───────────────────────────────────────────────────────────────

_LABELS = ("Código", "Fornecedor", "CNPJ", "Data Cobrança",
           "Data Vencimento", "Data Pagamento", "Status", "Valor (R$)")


def test_group_charges_groups_by_codigo_and_sums_valor():
    df = pd.DataFrame([
        {"Código": "PAG-1", "Fornecedor": "A", "CNPJ": "111", "Data Cobrança": "01/06/2026",
         "Data Vencimento": "10/06/2026", "Data Pagamento": "", "Status": "Pendente", "Valor (R$)": 50.0},
        {"Código": "PAG-1", "Fornecedor": "A", "CNPJ": "111", "Data Cobrança": "01/06/2026",
         "Data Vencimento": "10/06/2026", "Data Pagamento": "", "Status": "Pendente", "Valor (R$)": 20.0},
        {"Código": "PAG-2", "Fornecedor": "B", "CNPJ": "222", "Data Cobrança": "02/06/2026",
         "Data Vencimento": "12/06/2026", "Data Pagamento": "", "Status": "Pago", "Valor (R$)": 30.0},
    ])
    groups = ch.group_charges(df, *_LABELS)
    by_cod = {g["cod"]: g for g in groups}

    assert set(by_cod) == {"PAG-1", "PAG-2"}
    assert by_cod["PAG-1"]["n_itens"] == 2
    assert by_cod["PAG-1"]["valor_total"] == pytest.approx(70.0)
    assert by_cod["PAG-1"]["fornecedor"] == "A"
    assert by_cod["PAG-2"]["n_itens"] == 1
    assert by_cod["PAG-2"]["valor_total"] == pytest.approx(30.0)


def test_group_charges_missing_cod_column_returns_empty_list():
    df = pd.DataFrame({"Fornecedor": ["A"]})
    assert ch.group_charges(df, *_LABELS) == []


def test_group_charges_empty_dataframe_returns_empty_list():
    df = pd.DataFrame(columns=list(_LABELS))
    assert ch.group_charges(df, *_LABELS) == []


# ── generate_single_charge_xlsx_bytes ───────────────────────────────────────────

_INSERT_SQL = (
    "INSERT INTO historico_cobrancas "
    '(COD_LANCAMENTO, DATA_COBRANCA, DATA_VENCIMENTO, DATA_PAGAMENTO, '
    'CNPJ_FORNECEDOR, STATUS_COBRANCA, "ORDEM MESTRE", '
    '"DATA DE PRODUÇÃO ACABAMENTO", "FORNECEDOR", "QUANTIDADE", '
    '"REMONTE", "REAL CORTADO", "MINUTOS GERADOS", "VALOR DO PROCESSO BRL") '
    "VALUES (:cod, :data_cobranca, :data_vencimento, :data_pagamento, :cnpj, "
    ":status, :om, :data_producao, :fornecedor, :qtd, :remonte, :real_cortado, "
    ":minutos, :valor)"
)


@pytest.fixture
def temp_db(tmp_path, monkeypatch):
    """
    Isola os módulos de banco/histórico num SQLite temporário para o teste,
    usando a mesma camada SQLAlchemy da aplicação (só troca a DATABASE_URL
    para um arquivo local em vez do Postgres/Supabase real). Os SQLs do
    projeto usam apenas recursos suportados por ambos os dialetos
    (identificadores entre aspas duplas, placeholders nomeados via `text()`).
    """
    db_path = tmp_path / "test_analise_gal.db"
    sqlite_url = f"sqlite:///{db_path}"
    monkeypatch.setenv("DATABASE_URL", sqlite_url)
    # st.secrets tem prioridade sobre a variável de ambiente em _database_url();
    # força o teste a usar o SQLite temporário independentemente do que estiver
    # configurado em .streamlit/secrets.toml.
    monkeypatch.setattr(db, "_database_url", lambda: sqlite_url)
    db.get_engine.clear()
    db.create_tables.clear()
    ch.create_tables()
    yield db_path
    db.get_engine.clear()
    db.create_tables.clear()


def _insert_item(db_path, cod, cnpj, supplier, status, om, qtd, remonte, real_cortado, minutos, valor):
    with db.get_connection() as conn:
        conn.execute(
            text(_INSERT_SQL),
            {
                "cod": cod, "data_cobranca": "01/06/2026", "data_vencimento": "15/06/2026",
                "data_pagamento": "", "cnpj": cnpj, "status": status, "om": om,
                "data_producao": "01/06/2026", "fornecedor": supplier, "qtd": qtd,
                "remonte": remonte, "real_cortado": real_cortado, "minutos": minutos,
                "valor": valor,
            },
        )
        conn.commit()


def test_generate_single_charge_xlsx_bytes_returns_none_for_missing_code(temp_db):
    assert ch.generate_single_charge_xlsx_bytes("PAG-NAOEXISTE") is None


def test_generate_single_charge_xlsx_bytes_returns_none_on_empty_database(temp_db):
    # Equivalente ao antigo "banco inexistente" da era SQLite: no modelo
    # Postgres/SQLAlchemy, create_tables() sempre garante o schema, então a
    # base "vazia" (sem nenhum lançamento) é o caso análogo — deve retornar
    # None em vez de lançar exceção.
    assert ch.generate_single_charge_xlsx_bytes("PAG-QUALQUER") is None


def test_generate_single_charge_xlsx_bytes_scopes_to_one_charge(temp_db):
    _insert_item(temp_db, "PAG-AAAA1111", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)
    _insert_item(temp_db, "PAG-AAAA1111", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "101", 5, "SEM ARREMATE", 5, 2.0, 20.0)
    _insert_item(temp_db, "PAG-BBBB2222", "22.222.222/0001-22", "Fornecedor B",
                 "Pago", "200", 3, "TROCAR", 3, 1.0, 30.0)

    xlsx_bytes = ch.generate_single_charge_xlsx_bytes("PAG-AAAA1111")
    assert xlsx_bytes is not None

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    all_codes_in_sheet = [
        ws.cell(row=r, column=1).value
        for r in range(ch._HEADER_OFFSET + 1, ws.max_row + 1)
    ]
    # Só os 2 itens do lançamento pedido devem aparecer (nada do outro código).
    assert all_codes_in_sheet.count("PAG-AAAA1111") == 2
    assert "PAG-BBBB2222" not in all_codes_in_sheet


# ── STATUS_OPTIONS ──────────────────────────────────────────────────────────────

def test_status_options_uses_devolucao_not_contestado():
    assert ch.STATUS_OPTIONS == ["Pendente", "Pago", "Devolução"]
    assert "Contestado" not in ch.STATUS_OPTIONS


# ── update_lancamento_status: fluxo de Devolução ────────────────────────────────

def test_update_lancamento_status_devolucao_moves_row_to_devolucoes(temp_db):
    _insert_item(temp_db, "PAG-DEV0001", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)

    ok = ch.update_lancamento_status("PAG-DEV0001", "Devolução")
    assert ok is True

    with db.get_connection() as conn:
        remaining = conn.execute(
            text('SELECT COUNT(*) FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :cod'),
            {"cod": "PAG-DEV0001"},
        ).scalar()
        moved = conn.execute(
            text('SELECT "STATUS_COBRANCA", "DATA_PAGAMENTO" FROM devolucoes WHERE "COD_LANCAMENTO" = :cod'),
            {"cod": "PAG-DEV0001"},
        ).fetchone()

    assert remaining == 0
    assert moved is not None
    assert moved[0] == "Devolução"
    assert moved[1] == ""


def test_update_lancamento_status_devolucao_returns_false_for_unknown_code(temp_db):
    assert ch.update_lancamento_status("PAG-NAOEXISTE", "Devolução") is False


def test_update_lancamento_status_rejects_status_outside_status_options(temp_db):
    _insert_item(temp_db, "PAG-DEV0002", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)
    # "Contestado" não é mais uma opção válida de status.
    assert ch.update_lancamento_status("PAG-DEV0002", "Contestado") is False


# ── update_lancamento_status: fluxo de Pago e reivindicação ────────────────────

def test_update_lancamento_status_pago_moves_row_with_payment_date(temp_db):
    _insert_item(temp_db, "PAG-PAY0001", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)

    assert ch.update_lancamento_status(
        "PAG-PAY0001", "Pago", data_pagamento=date(2026, 6, 12)
    ) is True

    with db.get_connection() as conn:
        remaining = conn.execute(
            text('SELECT COUNT(*) FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :cod'),
            {"cod": "PAG-PAY0001"},
        ).scalar()
        moved = conn.execute(
            text(
                'SELECT "STATUS_COBRANCA", "DATA_PAGAMENTO", "VALOR DO PROCESSO BRL" '
                'FROM pagamentos_concluidos WHERE "COD_LANCAMENTO" = :cod'
            ),
            {"cod": "PAG-PAY0001"},
        ).fetchone()

    assert remaining == 0
    assert moved == ("Pago", "12/06/2026", 50.0)


def test_update_lancamento_status_second_call_returns_false_without_duplicating(temp_db):
    """
    Reivindicação: dois admins concluindo o mesmo lançamento. O segundo não pode
    gravar um pagamento duplicado — só o primeiro reivindica as linhas.
    """
    _insert_item(temp_db, "PAG-DUP0001", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)

    assert ch.update_lancamento_status(
        "PAG-DUP0001", "Pago", data_pagamento=date(2026, 6, 12)
    ) is True
    assert ch.update_lancamento_status(
        "PAG-DUP0001", "Pago", data_pagamento=date(2026, 6, 12)
    ) is False

    with db.get_connection() as conn:
        pagos = conn.execute(
            text('SELECT COUNT(*) FROM pagamentos_concluidos WHERE "COD_LANCAMENTO" = :cod'),
            {"cod": "PAG-DUP0001"},
        ).scalar()
    assert pagos == 1


def test_update_lancamento_status_propagates_database_error(temp_db, monkeypatch):
    """
    Uma falha de banco não pode virar `False`: isso a tornaria indistinguível de
    "código não encontrado", e o admin repetiria a ação sem saber que o Supabase
    é que estava fora. Deve propagar para a fronteira @page_guard da página.
    """
    _insert_item(temp_db, "PAG-ERR0001", "11.111.111/0001-11", "Fornecedor A",
                 "Pendente", "100", 10, "PONTO ESTOURADO", 10, 5.0, 50.0)

    def _boom():
        raise db.DatabaseUnavailableError("banco indisponível")

    monkeypatch.setattr(ch, "get_connection", _boom)

    with pytest.raises(db.DatabaseUnavailableError):
        ch.update_lancamento_status("PAG-ERR0001", "Pago", data_pagamento=date(2026, 6, 12))


# ── launch_charge: lançamento atômico com reivindicação ────────────────────────

_REGISTRO_SQL = (
    "INSERT INTO registros_defeitos "
    '("ORDEM MESTRE", "DATA DE PRODUÇÃO ACABAMENTO", "FORNECEDOR", "QUANTIDADE", '
    '"REMONTE", "REAL CORTADO", "MINUTOS GERADOS", "VALOR DO PROCESSO BRL", '
    '"STATUS_COBRANCA") '
    "VALUES (:om, :dp, :sup, :qtd, :rem, :rc, :mins, :val, 'Pendente')"
)


def _insert_registro(supplier, data_producao="2026-07-10"):
    with db.get_connection() as conn:
        conn.execute(
            text(_REGISTRO_SQL),
            {"om": "OM-1", "dp": data_producao, "sup": supplier, "qtd": 2,
             "rem": "PONTO ESTOURADO", "rc": "10", "mins": 5.0, "val": 100.0},
        )
        conn.commit()


def _charge_df(supplier, valor=100.0, qtd=2):
    return pd.DataFrame([{
        COLS["order"]:     "OM-1",
        COLS["date"]:      pd.Timestamp("2026-07-10"),
        COLS["supplier"]:  supplier,
        COLS["quantity"]:  qtd,
        COLS["defect"]:    "PONTO ESTOURADO",
        COLS["real_cut"]:  "10",
        COLS["minutes"]:   5.0,
        COLS["value_brl"]: valor,
    }])


def _launch(supplier, **kwargs):
    return ch.launch_charge(
        supplier=supplier,
        cnpj="00.000.000/0001-00",
        df_records=_charge_df(supplier),
        data_cobranca=date(2026, 7, 16),
        data_vencimento=date(2026, 7, 26),
        reference_date=date(2026, 7, 10),
        reference_date_end=date(2026, 7, 10),
        **kwargs,
    )


def test_launch_charge_claims_records_and_writes_charge(temp_db):
    _insert_registro("OFICINA X")

    cod = _launch("OFICINA X")

    with db.get_connection() as conn:
        restantes = conn.execute(
            text('SELECT COUNT(*) FROM registros_defeitos WHERE "FORNECEDOR" = :s'),
            {"s": "OFICINA X"},
        ).scalar()
        cobrados = conn.execute(
            text('SELECT COUNT(*) FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :c'),
            {"c": cod},
        ).scalar()

    # Os registros saíram da base ativa e a cobrança ficou gravada — na mesma transação.
    assert restantes == 0
    assert cobrados == 1


def test_launch_charge_second_launch_raises_and_writes_nothing(temp_db):
    """
    Regressão da cobrança duplicada: dois admins com a página aberta lançam a
    mesma cobrança. O segundo não encontra registros a reivindicar e deve falhar
    sem gravar — senão o fornecedor receberia duas cobranças do mesmo período.
    """
    _insert_registro("OFICINA Y")
    _launch("OFICINA Y")

    with pytest.raises(ch.ChargeAlreadyLaunchedError):
        _launch("OFICINA Y")

    with db.get_connection() as conn:
        lancamentos = conn.execute(
            text(
                'SELECT COUNT(DISTINCT "COD_LANCAMENTO") FROM historico_cobrancas '
                'WHERE "FORNECEDOR" = :s'
            ),
            {"s": "OFICINA Y"},
        ).scalar()
    assert lancamentos == 1


def test_launch_charge_outside_reference_period_does_not_claim(temp_db):
    # Registro de outra data não pode ser reivindicado pelo período selecionado.
    _insert_registro("OFICINA Z", data_producao="2026-08-01")

    with pytest.raises(ch.ChargeAlreadyLaunchedError):
        _launch("OFICINA Z")

    with db.get_connection() as conn:
        restantes = conn.execute(
            text('SELECT COUNT(*) FROM registros_defeitos WHERE "FORNECEDOR" = :s'),
            {"s": "OFICINA Z"},
        ).scalar()
        cobrados = conn.execute(
            text('SELECT COUNT(*) FROM historico_cobrancas WHERE "FORNECEDOR" = :s'),
            {"s": "OFICINA Z"},
        ).scalar()
    assert restantes == 1   # nada foi apagado
    assert cobrados == 0    # nada foi gravado


def test_launch_charge_split_writes_both_halves_with_same_code(temp_db):
    import src.data.divida_dividida as dd

    _insert_registro("OFICINA W")

    cod = _launch("OFICINA W", df_empresa=_charge_df("OFICINA W", valor=40.0))

    with db.get_connection() as conn:
        forn = conn.execute(
            text(
                'SELECT "VALOR DO PROCESSO BRL" FROM historico_cobrancas '
                'WHERE "COD_LANCAMENTO" = :c'
            ),
            {"c": cod},
        ).fetchone()
        emp = conn.execute(
            text(
                'SELECT "VALOR DO PROCESSO BRL", "STATUS_COBRANCA" '
                'FROM tb_divida_dividida WHERE "COD_LANCAMENTO" = :c'
            ),
            {"c": cod},
        ).fetchone()

    assert forn[0] == 100.0
    assert emp[0] == 40.0
    assert emp[1] == dd.STATUS_DIVIDIDA


# ── migrate_contestado_to_devolucao: compatibilidade com dados legados ─────────

def test_migrate_contestado_to_devolucao_moves_legacy_rows(temp_db):
    _insert_item(temp_db, "PAG-LEG0001", "22.222.222/0001-22", "Fornecedor B",
                 "Contestado", "200", 3, "TROCAR", 3, 1.0, 30.0)

    moved = ch.migrate_contestado_to_devolucao()
    assert moved == 1

    with db.get_connection() as conn:
        remaining = conn.execute(
            text('SELECT COUNT(*) FROM historico_cobrancas WHERE "STATUS_COBRANCA" = \'Contestado\'')
        ).scalar()
        dev_row = conn.execute(
            text('SELECT "STATUS_COBRANCA" FROM devolucoes WHERE "COD_LANCAMENTO" = :cod'),
            {"cod": "PAG-LEG0001"},
        ).fetchone()

    assert remaining == 0
    assert dev_row is not None
    assert dev_row[0] == "Devolução"


def test_migrate_contestado_to_devolucao_is_idempotent(temp_db):
    _insert_item(temp_db, "PAG-LEG0002", "22.222.222/0001-22", "Fornecedor B",
                 "Contestado", "200", 3, "TROCAR", 3, 1.0, 30.0)

    first  = ch.migrate_contestado_to_devolucao()
    second = ch.migrate_contestado_to_devolucao()

    assert first == 1
    assert second == 0


def test_migrate_contestado_to_devolucao_noop_when_no_legacy_rows(temp_db):
    assert ch.migrate_contestado_to_devolucao() == 0
