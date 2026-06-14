# -*- coding: utf-8 -*-
"""
Gerador de documento de cobrança em Excel.
Cria um arquivo .xlsx formatado profissionalmente para envio ao fornecedor.
"""

import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paleta de cores (hex sem #) ────────────────────────────────────────────────
_PURPLE_DARK  = "1A1530"   # fundo cabeçalho principal
_PURPLE_MID   = "534AB7"   # fundo cabeçalho da tabela
_PURPLE_LIGHT = "EDE8FF"   # fundo linhas alternadas
_RED_ALERT    = "C0392B"   # total destaque
_WHITE        = "FFFFFF"
_GRAY_LIGHT   = "F5F5F5"
_GRAY_BORDER  = "CCCCCC"
_BLACK        = "000000"


def _thin_border(color: str = _GRAY_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_border() -> Border:
    thick = Side(style="medium", color=_PURPLE_MID)
    thin  = Side(style="thin",   color=_PURPLE_MID)
    return Border(left=thin, right=thin, top=thick, bottom=thick)


def generate_charge_excel(
    supplier: str,
    cnpj: str,
    df_records: pd.DataFrame,
    display_cols: list[str],
    col_labels: dict[str, str],
) -> bytes:
    """
    Gera documento de cobrança em Excel para um fornecedor.

    O CNPJ do fornecedor é exibido no cabeçalho do documento.
    A tabela de registros contém apenas as colunas de defeito/processo.

    Args:
        supplier:     Nome do fornecedor.
        cnpj:         CNPJ do fornecedor (string formatada, ex: XX.XXX.XXX/XXXX-XX).
        df_records:   DataFrame com os registros de defeito do fornecedor.
        display_cols: Lista de colunas a incluir na tabela (nomes internos).
        col_labels:   Mapeamento {coluna_interna: label_exibição}.

    Returns:
        Bytes do arquivo .xlsx pronto para download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobrança"

    today    = date.today().strftime("%d/%m/%Y")
    num_cols = len(display_cols)
    last_col = get_column_letter(num_cols)

    # ── 1. CABEÇALHO DO DOCUMENTO ─────────────────────────────────────────────

    # Linha 1: Título principal
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "AVISO DE COBRANÇA — DEFEITOS / REMONTES"
    c.font  = Font(name="Calibri", bold=True, size=15, color=_WHITE)
    c.fill  = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Linha 2: Data de emissão
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Emitido em: {today}"
    c.font  = Font(name="Calibri", size=10, color="C8C0F0")
    c.fill  = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Linha 3: espaço
    ws.row_dimensions[3].height = 8

    # Linhas 4-5: Dados do fornecedor (CNPJ informado pelo usuário no cabeçalho)
    _write_info_row(ws, 4, "Fornecedor:", supplier, num_cols)
    _write_info_row(ws, 5, "CNPJ:",       cnpj,     num_cols)

    # Linha 6: espaço
    ws.row_dimensions[6].height = 8

    # ── 2. CABEÇALHO DA TABELA ────────────────────────────────────────────────
    header_row = 7
    for idx, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=header_row, column=idx)
        cell.value     = col_labels.get(col, col)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _header_border()
    ws.row_dimensions[header_row].height = 28

    # ── 3. DADOS ──────────────────────────────────────────────────────────────
    total_value    = 0.0
    value_col_name = "VALOR DO PROCESSO BRL"

    for row_idx, (_, row) in enumerate(df_records.iterrows(), start=header_row + 1):
        fill_color = _PURPLE_LIGHT if row_idx % 2 == 0 else _WHITE

        for col_idx, col in enumerate(display_cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _thin_border()
            cell.font   = Font(name="Calibri", size=9)
            cell.fill   = PatternFill("solid", fgColor=fill_color)

            val = row.get(col)
            if pd.isna(val):
                val = ""

            if col == "DATA DE PRODUÇÃO ACABAMENTO":
                if hasattr(val, "strftime"):
                    val = val.strftime("%d/%m/%Y")
                cell.alignment = Alignment(horizontal="center")
            elif col == value_col_name:
                total_value += float(val) if val != "" else 0
                cell.number_format = 'R$ #,##0.00'
                cell.alignment     = Alignment(horizontal="right")
                cell.value         = float(val) if val != "" else 0
                continue
            elif col in ("QUANTIDADE", "MINUTOS GERADOS", "REAL CORTADO"):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

            cell.value = val

        ws.row_dimensions[row_idx].height = 18

    # ── 4. LINHA DE TOTAL ─────────────────────────────────────────────────────
    total_row     = header_row + len(df_records) + 1
    value_col_idx = (
        display_cols.index(value_col_name) + 1
        if value_col_name in display_cols
        else len(display_cols)
    )
    merge_end = get_column_letter(value_col_idx - 1)
    ws.merge_cells(f"A{total_row}:{merge_end}{total_row}")

    label_cell = ws[f"A{total_row}"]
    label_cell.value     = "TOTAL A COBRAR"
    label_cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    label_cell.fill      = PatternFill("solid", fgColor=_RED_ALERT)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border    = _thin_border(_RED_ALERT)

    total_cell = ws.cell(row=total_row, column=value_col_idx)
    total_cell.value         = total_value
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.font          = Font(name="Calibri", bold=True, size=11, color=_WHITE)
    total_cell.fill          = PatternFill("solid", fgColor=_RED_ALERT)
    total_cell.alignment     = Alignment(horizontal="right", vertical="center")
    total_cell.border        = _thin_border(_RED_ALERT)
    ws.row_dimensions[total_row].height = 22

    # Preencher colunas após o valor na linha de total
    for col_idx in range(value_col_idx + 1, len(display_cols) + 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=_RED_ALERT)
        c.border = _thin_border(_RED_ALERT)

    # ── 5. RODAPÉ ─────────────────────────────────────────────────────────────
    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = (
        "Este documento é gerado automaticamente pelo sistema de Controle de Qualidade. "
        "Favor providenciar o ressarcimento no prazo estipulado em contrato."
    )
    c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 30

    # ── 6. LARGURA DAS COLUNAS ────────────────────────────────────────────────
    col_widths = {
        "DATA DE PRODUÇÃO ACABAMENTO": 14,
        "FORNECEDOR":                  32,
        "ORDEM MESTRE":                16,
        "QUANTIDADE":                  12,
        "REMONTE":                     22,
        "REAL CORTADO":                14,
        "MINUTOS GERADOS":             16,
        "VALOR DO PROCESSO BRL":       18,
    }
    for idx, col in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col, 15)

    # ── Salvar em memória ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Private helpers ───────────────────────────────────────────────────────────

def _write_info_row(
    ws,
    row: int,
    label: str,
    value: str,
    num_cols: int,
) -> None:
    """Escreve uma linha de informação estilo 'Label: Valor'."""
    mid      = num_cols // 2
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f"A{row}:{get_column_letter(mid)}{row}")
    lc = ws[f"A{row}"]
    lc.value     = label
    lc.font      = Font(name="Calibri", bold=True, size=10, color=_PURPLE_MID)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18

    ws.merge_cells(f"{get_column_letter(mid + 1)}{row}:{last_col}{row}")
    vc = ws.cell(row=row, column=mid + 1)
    vc.value     = value
    vc.font      = Font(name="Calibri", bold=True, size=10, color=_BLACK)
    vc.alignment = Alignment(horizontal="left", vertical="center")
