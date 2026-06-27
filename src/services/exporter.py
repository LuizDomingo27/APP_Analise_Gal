"""
Exportador de dados agrupados por Fornecedor.
Tema clean: fundo branco, acentos #00805C / #00B884, linhas pares #F2F7F5.
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config.settings import COLS

# ── Paleta clean ──────────────────────────────────────────────────────────────
_DARK_TEAL  = "014B43"   # cabeçalho / título
_MID_TEAL   = "099078"   # totais / acento
_CREAM      = "F9ECE5"   # linha par
_WHITE      = "FFFFFF"   # linha ímpar / fundo geral
_TEXT_DARK  = "1A2E2A"   # texto principal
_TEXT_MID   = "4A7570"   # texto secundário
_BORDER_CLR = "D6E8E5"   # borda suave


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = _TEXT_DARK, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", color=hex_color, bold=bold, size=size)


def _border() -> Border:
    s = Side(style="thin", color=_BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


# ── Agregação ─────────────────────────────────────────────────────────────────

def _aggregate(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby(COLS["supplier"])

    def top_defect(sub):
        return sub[COLS["defect"]].value_counts().index[0] if len(sub) else "—"

    result = pd.DataFrame()
    result["Fornecedor"]         = g[COLS["supplier"]].first()
    result["Data Início"]        = g[COLS["date"]].min().dt.strftime("%d/%m/%Y")
    result["Data Fim"]           = g[COLS["date"]].max().dt.strftime("%d/%m/%Y")
    result["OM (únicos)"]        = g[COLS["order"]].nunique()
    result["Total de Ordem"]     = g[COLS["order"]].count()
    result["Defeito Principal"]  = df.groupby(COLS["supplier"]).apply(top_defect, include_groups=False)
    result["Total Defeitos"]     = g[COLS["defect"]].count()
    result["Total Peças"]        = g[COLS["quantity"]].sum()
    result["Total Minutos"]      = g[COLS["minutes"]].sum().round(2)
    result["Percentual (%)"]     = (g[COLS["pct_remonte"]].mean() * 100).round(2)
    result["Valor em Real (R$)"] = g[COLS["value_brl"]].sum().round(2)
    return result.reset_index(drop=True)


# ── Escrita Excel ─────────────────────────────────────────────────────────────

def _write_excel(df_agg: pd.DataFrame, ts: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo por Fornecedor"

    headers = list(df_agg.columns)
    n_cols  = len(headers)

    # ── Faixa de título ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1)
    tc.value     = f"Resumo de Defeitos por Fornecedor  ·  {ts}"
    tc.font      = _font(_WHITE, bold=True, size=12)
    tc.fill      = _fill(_DARK_TEAL)
    tc.alignment = _center()
    ws.row_dimensions[1].height = 28

    # ── Cabeçalho das colunas ─────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = header.upper()
        cell.font      = _font(_WHITE, bold=True, size=10)
        cell.fill      = _fill(_DARK_TEAL)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[2].height = 22

    # ── Linha de subtítulo com total de fornecedores ──────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    sc = ws.cell(row=3, column=1)
    sc.value     = f"Total de fornecedores: {len(df_agg)}"
    sc.font      = _font(_WHITE, size=9)
    sc.fill      = _fill(_MID_TEAL)
    sc.alignment = _center()
    ws.row_dimensions[3].height = 15

    # ── Dados ─────────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df_agg.iterrows(), start=4):
        bg = _CREAM if row_idx % 2 == 0 else _WHITE
        for col_idx, col_name in enumerate(headers, start=1):
            cell       = ws.cell(row=row_idx, column=col_idx)
            val        = row[col_name]
            cell.fill  = _fill(bg)
            cell.font  = _font(_TEXT_DARK)
            cell.border = _border()

            if col_name in ("Fornecedor", "Defeito Principal"):
                cell.alignment = _left()
            else:
                cell.alignment = _center()

            # Formato e valor
            if col_name == "Valor em Real (R$)":
                cell.value          = val
                cell.number_format  = '"R$" #,##0.00'
            elif col_name == "Total Minutos":
                cell.value          = val
                cell.number_format  = "#,##0.00"
            elif col_name == "Percentual (%)":
                cell.value          = val / 100 if isinstance(val, (int, float)) else val
                cell.number_format  = "0.00%"
            elif col_name in ("Total Peças", "Total Defeitos", "OM (únicos)", "Total de Ordem"):
                cell.value          = val
                cell.number_format  = "#,##0"
            else:
                cell.value = val
        ws.row_dimensions[row_idx].height = 18

    # ── Linha de totais ───────────────────────────────────────────────────────
    total_row = len(df_agg) + 4
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    lbl            = ws.cell(row=total_row, column=1)
    lbl.value      = "TOTAL GERAL"
    lbl.font       = _font(_WHITE, bold=True)
    lbl.fill       = _fill(_MID_TEAL)
    lbl.alignment  = _center()
    lbl.border     = _border()

    totals_map = {
        "OM (únicos)":        df_agg["OM (únicos)"].sum(),
        "Total de Ordem":     df_agg["Total de Ordem"].sum(),
        "Total Defeitos":     df_agg["Total Defeitos"].sum(),
        "Total Peças":        df_agg["Total Peças"].sum(),
        "Total Minutos":      round(df_agg["Total Minutos"].sum(), 2),
        "Valor em Real (R$)": round(df_agg["Valor em Real (R$)"].sum(), 2),
    }

    for col_idx, col_name in enumerate(headers, start=1):
        if col_idx <= 3:
            continue
        cell           = ws.cell(row=total_row, column=col_idx)
        val            = totals_map.get(col_name, "")
        cell.font      = _font(_WHITE, bold=True)
        cell.fill      = _fill(_MID_TEAL)
        cell.alignment = _center()
        cell.border    = _border()
        if col_name == "Valor em Real (R$)":
            cell.value, cell.number_format = val, '"R$" #,##0.00'
        elif col_name == "Total Minutos":
            cell.value, cell.number_format = val, "#,##0.00"
        elif col_name in ("Total Peças", "Total Defeitos", "OM (únicos)", "Total de Ordem"):
            cell.value, cell.number_format = val, "#,##0"
        else:
            cell.value = val
    ws.row_dimensions[total_row].height = 22

    # ── Larguras ──────────────────────────────────────────────────────────────
    widths = {
        "Fornecedor": 26, "Data Início": 14, "Data Fim": 14,
        "OM (únicos)": 13, "Total de Ordem": 15, "Defeito Principal": 22,
        "Total Defeitos": 15, "Total Peças": 13, "Total Minutos": 15,
        "Percentual (%)": 15, "Valor em Real (R$)": 20,
    }
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 14)

    ws.freeze_panes = "A4"

    # ── Aba Resumo (fornecedores com valor > R$ 100) ──────────────────────────
    _write_resumo_sheet(wb, df_agg, ts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_resumo_sheet(wb: Workbook, df_agg: pd.DataFrame, ts: str) -> None:
    """Cria a aba 'Resumo' com os fornecedores cujo Valor em Real > R$ 100."""

    resumo_cols = [
        "Fornecedor",
        "OM (únicos)",
        "Total Defeitos",
        "Total Minutos",
        "Total Peças",
        "Valor em Real (R$)",
        "Percentual (%)",
    ]

    df_resumo = df_agg[df_agg["Valor em Real (R$)"] > 100][resumo_cols].copy()
    df_resumo = df_resumo.sort_values("Valor em Real (R$)", ascending=False).reset_index(drop=True)

    ws = wb.create_sheet("Resumo")
    n_cols = len(resumo_cols)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1)
    tc.value     = f"Resumo de Fornecedores com Valor > R$ 100  ·  {ts}"
    tc.font      = _font(_WHITE, bold=True, size=12)
    tc.fill      = _fill(_DARK_TEAL)
    tc.alignment = _center()
    ws.row_dimensions[1].height = 28

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(resumo_cols, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = header.upper()
        cell.font      = _font(_WHITE, bold=True, size=10)
        cell.fill      = _fill(_DARK_TEAL)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[2].height = 22

    # ── Subtítulo ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    sc = ws.cell(row=3, column=1)
    sc.value     = f"{len(df_resumo)} fornecedor(es) com valor total acima de R$ 100,00"
    sc.font      = _font(_WHITE, size=9)
    sc.fill      = _fill(_MID_TEAL)
    sc.alignment = _center()
    ws.row_dimensions[3].height = 15

    # ── Dados ─────────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df_resumo.iterrows(), start=4):
        bg = _CREAM if row_idx % 2 == 0 else _WHITE
        for col_idx, col_name in enumerate(resumo_cols, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx)
            val         = row[col_name]
            cell.fill   = _fill(bg)
            cell.font   = _font(_TEXT_DARK)
            cell.border = _border()

            if col_name == "Fornecedor":
                cell.alignment = _left()
            else:
                cell.alignment = _center()

            if col_name == "Valor em Real (R$)":
                cell.value         = val
                cell.number_format = '"R$" #,##0.00'
            elif col_name == "Total Minutos":
                cell.value         = val
                cell.number_format = "#,##0.00"
            elif col_name == "Percentual (%)":
                cell.value         = val / 100 if isinstance(val, (int, float)) else val
                cell.number_format = "0.00%"
            elif col_name in ("Total Peças", "Total Defeitos", "OM (únicos)"):
                cell.value         = val
                cell.number_format = "#,##0"
            else:
                cell.value = val
        ws.row_dimensions[row_idx].height = 18

    # ── Linha de totais ───────────────────────────────────────────────────────
    total_row = len(df_resumo) + 4
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    lbl           = ws.cell(row=total_row, column=1)
    lbl.value     = "TOTAL"
    lbl.font      = _font(_WHITE, bold=True)
    lbl.fill      = _fill(_MID_TEAL)
    lbl.alignment = _center()
    lbl.border    = _border()

    totals_map = {
        "OM (únicos)":        df_resumo["OM (únicos)"].sum(),
        "Total Defeitos":     df_resumo["Total Defeitos"].sum(),
        "Total Minutos":      round(df_resumo["Total Minutos"].sum(), 2),
        "Total Peças":        df_resumo["Total Peças"].sum(),
        "Valor em Real (R$)": round(df_resumo["Valor em Real (R$)"].sum(), 2),
    }
    for col_idx, col_name in enumerate(resumo_cols, start=1):
        if col_idx <= 2:
            continue
        cell           = ws.cell(row=total_row, column=col_idx)
        val            = totals_map.get(col_name, "")
        cell.font      = _font(_WHITE, bold=True)
        cell.fill      = _fill(_MID_TEAL)
        cell.alignment = _center()
        cell.border    = _border()
        if col_name == "Valor em Real (R$)":
            cell.value, cell.number_format = val, '"R$" #,##0.00'
        elif col_name == "Total Minutos":
            cell.value, cell.number_format = val, "#,##0.00"
        elif col_name in ("Total Peças", "Total Defeitos", "OM (únicos)"):
            cell.value, cell.number_format = val, "#,##0"
        else:
            cell.value = val
    ws.row_dimensions[total_row].height = 22

    # ── Larguras ──────────────────────────────────────────────────────────────
    widths_r = {
        "Fornecedor": 28, "OM (únicos)": 13, "Total Defeitos": 15,
        "Total Minutos": 16, "Total Peças": 13, "Valor em Real (R$)": 20, "Percentual (%)": 16,
    }
    for col_idx, col_name in enumerate(resumo_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths_r.get(col_name, 14)

    ws.freeze_panes = "A4"


def get_xlsx_bytes(filtered_df: pd.DataFrame) -> bytes:
    """Gera e retorna os bytes do Excel — sem gravar nada em disco."""
    ts     = datetime.now().strftime("%d/%m/%Y %H:%M")
    df_agg = _aggregate(filtered_df)
    return _write_excel(df_agg, ts)
