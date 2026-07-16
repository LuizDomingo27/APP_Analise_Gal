# -*- coding: utf-8 -*-
"""
Gerenciamento dos Pagamentos Concluídos — tabela pagamentos_concluidos (Postgres/Supabase).
"""

import io
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config.settings import CACHE_TTL_SECONDS, COLS, DATASET_DIR
from src.data.database import create_tables, get_connection
from src.data.cobranca_history import (
    HISTORY_LABELS,
    _COL_WIDTHS,
    payment_punctuality,
)

# Legado: mantido para imports externos que ainda referenciam este símbolo
BD_PAGAMENTOS = DATASET_DIR / "bd_pagamentos.xlsx"

_PURPLE_DARK  = "1A1530"
_PURPLE_MID   = "534AB7"
_PURPLE_LIGHT = "EDE8FF"
_WHITE        = "FFFFFF"
_GRAY_BORDER  = "C8C0F0"
_TEXT_LIGHT   = "C8C0F0"
_GREEN        = "1D9E75"
_RED          = "C0392B"
_AMBER        = "D8932E"

_HEADER_OFFSET = 5


def append_payments(df_rows: pd.DataFrame) -> None:
    """
    Acrescenta linhas (já com STATUS='Pago') em pagamentos_concluidos.
    Chamada por update_lancamento_status quando o status muda para 'Pago'
    — o movimento é feito atomicamente dentro da mesma conexão lá,
    então esta função é usada apenas por chamadas externas / migração.
    """
    create_tables()

    # `id` é PK auto-gerada; não reinserir uma id de origem se vier no df.
    df_rows = df_rows.drop(columns=["id"], errors="ignore")

    with get_connection() as conn:
        df_rows.to_sql("pagamentos_concluidos", conn, if_exists="append", index=False)
        conn.commit()

    load_payments.clear()


@st.cache_data(ttl=CACHE_TTL_SECONDS)
def load_payments() -> pd.DataFrame | None:
    """Carrega o histórico completo de pagamentos_concluidos. Retorna None se vazio."""
    create_tables()
    with get_connection() as conn:
        df = pd.read_sql("SELECT * FROM pagamentos_concluidos", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    return df


def generate_payments_xlsx_bytes() -> bytes | None:
    """
    Gera o xlsx executivo de pagamentos em memória e retorna os bytes.
    Retorna None se não houver dados.
    """
    create_tables()
    with get_connection() as conn:
        df = pd.read_sql("SELECT * FROM pagamentos_concluidos", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    buf = io.BytesIO()
    _write_payments_xlsx(df, buf)
    return buf.getvalue()


# ── Privado — escrita xlsx ────────────────────────────────────────────────────

def _write_payments_xlsx(df: pd.DataFrame, dest) -> None:
    """
    Grava df como xlsx executivo em dest (Path ou file-like / BytesIO).
    """
    _pag_parsed = pd.to_datetime(df.get("DATA_PAGAMENTO", ""), format="%d/%m/%Y", errors="coerce")
    df = df.assign(_ord=_pag_parsed).sort_values("_ord", ascending=False, na_position="last").drop(columns="_ord")
    df = df.reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos Concluídos"

    all_cols  = list(df.columns)
    num_cols  = len(all_cols)
    last_col  = get_column_letter(num_cols)
    today_br  = date.today().strftime("%d/%m/%Y")
    n_records = len(df)

    value_col_name  = COLS["value_brl"]
    status_col_name = COLS["status"]
    cod_col_name    = "COD_LANCAMENTO"

    total_value    = float(pd.to_numeric(df[value_col_name], errors="coerce").fillna(0).sum()) if value_col_name in df.columns else 0.0
    n_lancamentos  = df[cod_col_name].nunique() if cod_col_name in df.columns else n_records
    n_fornecedores = df[COLS["supplier"]].nunique() if COLS["supplier"] in df.columns else 0

    n_no_prazo = 0
    n_atraso   = 0
    if "DATA_PAGAMENTO" in df.columns and "DATA_VENCIMENTO" in df.columns:
        _seen = set()
        for _, r in df.iterrows():
            c = r.get(cod_col_name)
            if c in _seen:
                continue
            _seen.add(c)
            _, atrasado = payment_punctuality(r.get("DATA_PAGAMENTO"), r.get("DATA_VENCIMENTO"))
            if atrasado is True:
                n_atraso += 1
            elif atrasado is False:
                n_no_prazo += 1

    def thin(color=_GRAY_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_border():
        thick = Side(style="medium", color=_PURPLE_MID)
        thin_ = Side(style="thin",   color=_PURPLE_MID)
        return Border(left=thin_, right=thin_, top=thick, bottom=thick)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "PAGAMENTOS CONCLUÍDOS — RELATÓRIO EXECUTIVO"
    c.font      = Font(name="Calibri", bold=True, size=15, color=_WHITE)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"Gerado em: {today_br}  ·  Lançamentos pagos: {n_lancamentos}  ·  Itens: {n_records}"
    c.font      = Font(name="Calibri", size=10, color=_TEXT_LIGHT)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    kpis = [
        ("VALOR TOTAL PAGO", f"R$ {total_value:,.2f}", _GREEN),
        ("LANÇAMENTOS PAGOS", str(n_lancamentos), _PURPLE_MID),
        ("FORNECEDORES", str(n_fornecedores), _PURPLE_MID),
        ("NO PRAZO", str(n_no_prazo), _GREEN),
        ("COM ATRASO", str(n_atraso), _RED),
    ]
    n_kpi      = len(kpis)
    base_width = num_cols // n_kpi
    extra      = num_cols % n_kpi
    col_cursor = 1
    for i, (label, value, color) in enumerate(kpis):
        span       = base_width + (1 if i < extra else 0)
        span       = max(span, 1)
        start_col  = col_cursor
        end_col    = min(col_cursor + span - 1, num_cols)
        col_cursor = end_col + 1

        start_letter = get_column_letter(start_col)
        end_letter   = get_column_letter(max(end_col, start_col))
        ws.merge_cells(f"{start_letter}3:{end_letter}3")
        cell           = ws[f"{start_letter}3"]
        cell.value     = f"{label}:  {value}"
        cell.font      = Font(name="Calibri", bold=True, size=10, color=color)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_LIGHT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 6

    header_row = _HEADER_OFFSET
    col_labels = dict(HISTORY_LABELS)
    col_labels["COD_LANCAMENTO"] = "Código do Pagamento"

    for idx, col in enumerate(all_cols, start=1):
        cell           = ws.cell(row=header_row, column=idx)
        cell.value     = col_labels.get(col, col)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = header_border()
    ws.row_dimensions[header_row].height = 26

    last_cod    = None
    band_toggle = False

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        cur_cod = row.get(cod_col_name)
        if cur_cod != last_cod:
            band_toggle = not band_toggle
            last_cod = cur_cod
        fill_color = _PURPLE_LIGHT if band_toggle else _WHITE

        for col_idx, col in enumerate(all_cols, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin()

            val = row.get(col)
            if pd.isna(val):
                val = ""

            if col == status_col_name:
                cell.value     = "Pago"
                cell.fill      = PatternFill("solid", fgColor=_GREEN)
                cell.font      = Font(name="Calibri", size=9, bold=True, color=_WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            if col == value_col_name:
                fval               = float(val) if val != "" else 0.0
                cell.value         = fval
                cell.number_format = 'R$ #,##0.00'
                cell.fill          = PatternFill("solid", fgColor=fill_color)
                cell.alignment     = Alignment(horizontal="right")
                cell.font          = Font(name="Calibri", size=9, color="1A1530")
                continue

            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Calibri", size=9)

            if col == cod_col_name:
                cell.font      = Font(name="Consolas", size=9, bold=True, color=_PURPLE_MID)
                cell.alignment = Alignment(horizontal="center")
            elif col == "DATA_PAGAMENTO":
                cell.alignment = Alignment(horizontal="center")
                _dias_atraso, _atrasado = payment_punctuality(val, row.get("DATA_VENCIMENTO"))
                if _atrasado:
                    cell.font = Font(name="Calibri", size=9, bold=True, color=_RED)
                elif _atrasado is False:
                    cell.font = Font(name="Calibri", size=9, bold=True, color=_GREEN)
                else:
                    cell.font = Font(name="Calibri", size=9, italic=True, color=_AMBER)
            elif col in ("DATA_COBRANCA", "DATA_VENCIMENTO", COLS["date"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == "CNPJ_FORNECEDOR":
                cell.font      = Font(name="Calibri", size=9, color=_GREEN, bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col in (COLS["quantity"], COLS["real_cut"], COLS["minutes"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == COLS["order"]:
                cell.font      = Font(name="Calibri", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

            cell.value = val

        ws.row_dimensions[row_idx].height = 17

    total_row   = header_row + n_records + 1
    val_col_idx = (all_cols.index(value_col_name) + 1 if value_col_name in all_cols else num_cols)
    merge_end   = get_column_letter(max(val_col_idx - 1, 1))

    ws.merge_cells(f"A{total_row}:{merge_end}{total_row}")
    lc           = ws[f"A{total_row}"]
    lc.value     = "TOTAL PAGO"
    lc.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    lc.fill      = PatternFill("solid", fgColor=_GREEN)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border    = thin(_GREEN)

    tc               = ws.cell(row=total_row, column=val_col_idx)
    tc.value         = total_value
    tc.number_format = 'R$ #,##0.00'
    tc.font          = Font(name="Calibri", bold=True, size=11, color=_WHITE)
    tc.fill          = PatternFill("solid", fgColor=_GREEN)
    tc.alignment     = Alignment(horizontal="right", vertical="center")
    tc.border        = thin(_GREEN)
    ws.row_dimensions[total_row].height = 22

    for col_idx in range(val_col_idx + 1, num_cols + 1):
        c        = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=_GREEN)
        c.border = thin(_GREEN)

    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    c           = ws[f"A{footer_row}"]
    c.value     = (
        "Documento gerado automaticamente pelo sistema de Controle de Qualidade. "
        "Cada Código do Pagamento identifica um lançamento de cobrança pago integralmente."
    )
    c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 28

    for idx, col in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(col, 16)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(dest)
