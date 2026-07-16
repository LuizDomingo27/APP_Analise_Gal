# -*- coding: utf-8 -*-
"""
Gerenciamento do histórico de cobranças — tabela historico_cobrancas (Postgres/Supabase).
"""

import io
import logging
import uuid
import hashlib
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import text

from src.config.settings import CACHE_TTL_SECONDS, COLS, DATASET_DIR
from src.data.database import advisory_lock, create_tables, get_connection

logger = logging.getLogger(__name__)


class ChargeAlreadyLaunchedError(RuntimeError):
    """
    Os registros da cobrança já não estavam mais na base ativa no momento da
    gravação — outra sessão lançou essa mesma cobrança primeiro. Levantada por
    `launch_charge` com a transação já revertida: nada foi gravado.
    """

# ── Legado: mantido para imports externos que ainda referenciam este caminho ──
_BASE_DIR   = Path(__file__).resolve().parents[2]
BD_COBRANCA = DATASET_DIR / "bd_cobranca.xlsx"   # usado apenas como referência de nome

# ── Valores válidos de status ─────────────────────────────────────────────────
STATUS_OPTIONS = ["Pendente", "Pago", "Devolução"]
STATUS_DEFAULT = "Pendente"

# ── Paleta ────────────────────────────────────────────────────────────────────
_PURPLE_DARK  = "1A1530"
_PURPLE_MID   = "534AB7"
_PURPLE_LIGHT = "EDE8FF"
_WHITE        = "FFFFFF"
_GRAY_BORDER  = "C8C0F0"
_TEXT_LIGHT   = "C8C0F0"

_STATUS_COLORS = {
    "Pago":      {"bg": "1D9E75", "fg": "FFFFFF"},
    "Pendente":  {"bg": "EF9F27", "fg": "1A1530"},
    "Devolução": {"bg": "0F86A3", "fg": "FFFFFF"},
}

_HEADER_OFFSET = 4

_SAVE_COLS = [
    COLS["order"],
    COLS["date"],
    COLS["supplier"],
    COLS["quantity"],
    COLS["defect"],
    COLS["real_cut"],
    COLS["minutes"],
    COLS["value_brl"],
]

# Colunas de dados comuns a historico_cobrancas, pagamentos_concluidos e
# devolucoes — as três tabelas têm o mesmo schema. `id` fica de fora de
# propósito: é PK auto-gerada, e o destino do move gera a sua própria.
_CHARGE_COLUMNS = [
    "COD_LANCAMENTO",
    "DATA_COBRANCA",
    "DATA_VENCIMENTO",
    "DATA_PAGAMENTO",
    "CNPJ_FORNECEDOR",
    COLS["status"],
    COLS["order"],
    COLS["date"],
    COLS["supplier"],
    COLS["quantity"],
    COLS["defect"],
    COLS["real_cut"],
    COLS["minutes"],
    COLS["value_brl"],
]

# Status que tiram o lançamento do fluxo de cobrança, movendo-o para outra
# tabela. "Pendente" não está aqui: ele permanece em historico_cobrancas.
_TERMINAL_TABLES = {
    "Pago":      "pagamentos_concluidos",
    "Devolução": "devolucoes",
}

HISTORY_LABELS = {
    COLS["order"]:     "OM",
    COLS["date"]:      "Data Produção",
    COLS["supplier"]:  "Fornecedor",
    COLS["quantity"]:  "Qtd",
    COLS["defect"]:    "Remonte",
    COLS["real_cut"]:  "Real Cortado",
    COLS["minutes"]:   "Min. Gerados",
    COLS["value_brl"]: "Valor (R$)",
    "COD_LANCAMENTO":  "Código",
    "DATA_COBRANCA":   "Data Cobrança",
    "DATA_VENCIMENTO": "Data Vencimento",
    "DATA_PAGAMENTO":  "Data Pagamento",
    "CNPJ_FORNECEDOR": "CNPJ",
    COLS["status"]:    "Status",
}

_COL_WIDTHS = {
    "COD_LANCAMENTO":                   16,
    "DATA_COBRANCA":                    14,
    "DATA_VENCIMENTO":                  14,
    "DATA_PAGAMENTO":                   14,
    "CNPJ_FORNECEDOR":                  22,
    COLS["status"]:                     16,
    COLS["order"]:                      16,
    COLS["date"]:                       16,
    COLS["supplier"]:                   34,
    COLS["quantity"]:                   12,
    COLS["defect"]:                     26,
    COLS["real_cut"]:                   14,
    COLS["minutes"]:                    16,
    COLS["value_brl"]:                  20,
}


def payment_punctuality(data_pagamento, data_vencimento) -> tuple[int | None, bool | None]:
    """
    Compara Data Pagamento × Data Vencimento.
    Retorna (dias_de_atraso, atrasado):
      dias > 0  → pago com atraso
      dias <= 0 → no prazo
      (None, None) → faltam dados
    """
    venc = pd.to_datetime(data_vencimento, format="%d/%m/%Y", errors="coerce") \
        if isinstance(data_vencimento, str) else pd.to_datetime(data_vencimento, errors="coerce")
    pag = pd.to_datetime(data_pagamento, format="%d/%m/%Y", errors="coerce") \
        if isinstance(data_pagamento, str) else pd.to_datetime(data_pagamento, errors="coerce")

    if pag is None or pd.isna(pag) or venc is None or pd.isna(venc):
        return None, None

    dias = (pag.date() - venc.date()).days
    return dias, dias > 0


def gerar_cod_lancamento() -> str:
    return f"PAG-{uuid.uuid4().hex[:8].upper()}"


def status_badge_html(status: str) -> str:
    """Retorna o HTML de badge colorido (classe CSS badge-status) para o status informado."""
    s = str(status).strip()
    if s == "Pago":
        return '<span class="badge-status status-pago">✅ Pago</span>'
    if s == "Devolução":
        return '<span class="badge-status status-devolucao">🔄 Devolução</span>'
    return '<span class="badge-status status-pendente">⏳ Pendente</span>'


def situacao_badge_html(status: str, data_vencimento, data_pagamento) -> str:
    """
    Badge de situação do lançamento:
      - Pago -> compara Data de Pagamento x Data de Vencimento (pontualidade).
      - Pendente -> contagem regressiva até o vencimento.
    """
    s = str(status).strip()
    if s == "Pago":
        dias_atraso, atrasado = payment_punctuality(data_pagamento, data_vencimento)
        if atrasado is None:
            return '<span class="badge-status status-pendente">❔ Informe a data do pagamento</span>'
        if atrasado:
            return f'<span class="badge-status status-contestado">⚠️ Pago com {dias_atraso}d de atraso</span>'
        return '<span class="badge-status status-pago">✅ Pago no prazo</span>'

    venc_dt = pd.to_datetime(data_vencimento, format="%d/%m/%Y", errors="coerce") \
        if isinstance(data_vencimento, str) else pd.to_datetime(data_vencimento, errors="coerce")
    if venc_dt is None or pd.isna(venc_dt):
        return ""
    dias = (venc_dt.date() - date.today()).days
    if dias < 0:
        return f'<span class="badge-status status-contestado">⚠️ Vencido há {abs(dias)}d</span>'
    if dias == 0:
        return '<span class="badge-status status-pendente">⏳ Vence hoje</span>'
    return f'<span style="color:#00805C;font-weight:600">{dias} dia(s)</span>'


def group_charges(
    df: pd.DataFrame,
    cod_label: str,
    sup_label: str,
    cnpj_label: str,
    dte_label: str,
    venc_label: str,
    pag_label: str,
    status_label: str,
    val_label: str,
) -> list[dict]:
    """
    Agrupa um DataFrame de histórico já rotulado (uma linha por item) por
    Código de Lançamento, retornando um resumo — uma entrada por cobrança —
    com valor total e número de itens. Base para a tabela de extratos e para
    os seletores de "lançamento" usados na página de Histórico.
    """
    groups: list[dict] = []
    if cod_label not in df.columns:
        return groups
    for cod, grupo in df.groupby(cod_label, sort=False):
        primeira = grupo.iloc[0]
        valor_total = (
            pd.to_numeric(grupo[val_label], errors="coerce").sum()
            if val_label in grupo.columns else 0.0
        )
        groups.append({
            "cod": cod,
            "fornecedor": primeira.get(sup_label, ""),
            "cnpj": primeira.get(cnpj_label, ""),
            "data_cobranca": primeira.get(dte_label, ""),
            "data_vencimento": primeira.get(venc_label, ""),
            "data_pagamento": primeira.get(pag_label, ""),
            "status": primeira.get(status_label, ""),
            "n_itens": len(grupo),
            "valor_total": float(valor_total) if pd.notna(valor_total) else 0.0,
        })
    return groups


def _cod_lancamento_fallback(cnpj: str, data_cobranca: str) -> str:
    chave = f"{cnpj}|{data_cobranca}".encode("utf-8")
    return "LEG-" + hashlib.md5(chave).hexdigest()[:8].upper()


# ── Público: escrita ──────────────────────────────────────────────────────────

def build_charge_rows(
    df_records: pd.DataFrame,
    cod_lancamento: str,
    data_cobranca: date,
    data_vencimento: date,
    cnpj: str,
    status: str = STATUS_DEFAULT,
) -> pd.DataFrame:
    """
    Monta o DataFrame pronto para `to_sql` nas tabelas de cobrança
    (historico_cobrancas / tb_divida_dividida): seleciona as colunas de dados
    (_SAVE_COLS), formata datas para dd/mm/aaaa e insere as colunas de metadados
    (código, datas, CNPJ, status) nas posições esperadas pelo schema.

    Extraído de save_charge_to_history para ser reaproveitado pela gravação
    atômica da cobrança dividida (src/data/divida_dividida.py), evitando
    duplicar a lógica de montagem das linhas.
    """
    cols_to_save = [c for c in _SAVE_COLS if c in df_records.columns]
    df_save = df_records[cols_to_save].copy()

    for col in df_save.columns:
        if pd.api.types.is_datetime64_any_dtype(df_save[col]):
            df_save[col] = df_save[col].dt.strftime("%d/%m/%Y")

    df_save.insert(0, "COD_LANCAMENTO",  cod_lancamento)
    df_save.insert(1, "DATA_COBRANCA",   data_cobranca.strftime("%d/%m/%Y"))
    df_save.insert(2, "DATA_VENCIMENTO", data_vencimento.strftime("%d/%m/%Y"))
    df_save.insert(3, "DATA_PAGAMENTO",  "")
    df_save.insert(4, "CNPJ_FORNECEDOR", cnpj)
    df_save.insert(5, COLS["status"],    status)
    return df_save


def save_charge_to_history(
    supplier: str,
    cnpj: str,
    total: float,
    df_records: pd.DataFrame,
    display_cols: list[str],
    data_cobranca: date,
    data_vencimento: date,
    cod_lancamento: str | None = None,
) -> str:
    """
    Persiste os registros da cobrança na tabela historico_cobrancas (Postgres).
    Retorna o COD_LANCAMENTO usado.

    `cod_lancamento` pode ser informado pelo chamador para que a metade do
    fornecedor (aqui) e a metade da empresa (tb_divida_dividida) compartilhem o
    mesmo código numa cobrança dividida. Quando omitido, é gerado internamente
    (comportamento legado).
    """
    create_tables()

    if cod_lancamento is None:
        cod_lancamento = gerar_cod_lancamento()

    df_save = build_charge_rows(
        df_records, cod_lancamento, data_cobranca, data_vencimento, cnpj
    )

    with get_connection() as conn:
        insert_charge_rows(conn, df_save)
        conn.commit()

    load_history.clear()
    return cod_lancamento


def insert_charge_rows(
    conn, rows_fornecedor: pd.DataFrame, rows_empresa: pd.DataFrame | None = None
) -> None:
    """
    Insere as linhas de uma cobrança usando a conexão/transação de `conn`, sem
    comitar — quem chama decide o escopo da transação. `rows_empresa`, quando
    informado, é a metade absorvida pela empresa numa cobrança dividida e vai
    para tb_divida_dividida.

    Ponto único de escrita das cobranças: usado tanto pelo caminho simples
    (save_charge_to_history / save_split_charge) quanto pelo lançamento atômico
    com reivindicação (launch_charge), para que as três rotas gravem igual.
    """
    rows_fornecedor.to_sql("historico_cobrancas", conn, if_exists="append", index=False)
    if rows_empresa is not None:
        rows_empresa.to_sql("tb_divida_dividida", conn, if_exists="append", index=False)


def _claim_records(
    conn,
    supplier: str,
    reference_date: date | None,
    reference_date_end: date | None,
) -> int:
    """
    Reivindica os registros do fornecedor em registros_defeitos, apagando-os, e
    retorna quantos foram apagados. Zero significa que não havia nada a cobrar —
    outra sessão já lançou esta cobrança.

    É o DELETE que serve de trava: ele bloqueia as linhas, então duas sessões
    lançando a mesma cobrança ao mesmo tempo se serializam aqui, e a segunda
    encontra zero linhas depois que a primeira comita.
    """
    if reference_date is not None:
        date_end = reference_date_end if reference_date_end is not None else reference_date
        return conn.execute(
            text(
                'DELETE FROM registros_defeitos '
                'WHERE "FORNECEDOR" = :sup '
                'AND "DATA DE PRODUÇÃO ACABAMENTO" BETWEEN :d1 AND :d2'
            ),
            {
                "sup": supplier,
                "d1": reference_date.strftime("%Y-%m-%d"),
                "d2": date_end.strftime("%Y-%m-%d"),
            },
        ).rowcount
    return conn.execute(
        text('DELETE FROM registros_defeitos WHERE "FORNECEDOR" = :sup'),
        {"sup": supplier},
    ).rowcount


def launch_charge(
    supplier: str,
    cnpj: str,
    df_records: pd.DataFrame,
    data_cobranca: date,
    data_vencimento: date,
    reference_date: date | None = None,
    reference_date_end: date | None = None,
    df_empresa: pd.DataFrame | None = None,
) -> str:
    """
    Lança uma cobrança e retorna o COD_LANCAMENTO gerado. Numa ÚNICA transação:
    apaga os registros do fornecedor no período de registros_defeitos e grava a
    cobrança em historico_cobrancas (mais a metade da empresa em
    tb_divida_dividida, quando `df_empresa` é informado).

    Levanta ChargeAlreadyLaunchedError, sem gravar nada, se não havia registros
    a apagar: isso significa que outra sessão já lançou esta cobrança. Sem essa
    reivindicação, dois admins com a página aberta lançariam a mesma cobrança
    com dois códigos diferentes — cada um decide o que cobrar a partir do
    snapshot em `session_state`, que não sabe do que o outro fez.

    A transação única também garante que nunca fique cobrança gravada com os
    registros de origem ainda na base ativa (o que convidaria a uma segunda
    cobrança do mesmo período) nem o contrário.
    """
    create_tables()
    if df_empresa is not None:
        # Import local: divida_dividida importa deste módulo (ciclo no topo).
        from src.data.divida_dividida import STATUS_DIVIDIDA, _ensure_schema
        _ensure_schema()

    cod_lancamento = gerar_cod_lancamento()

    rows_forn = build_charge_rows(
        df_records, cod_lancamento, data_cobranca, data_vencimento, cnpj,
        status=STATUS_DEFAULT,
    )
    rows_emp = (
        build_charge_rows(
            df_empresa, cod_lancamento, data_cobranca, data_vencimento, cnpj,
            status=STATUS_DIVIDIDA,
        )
        if df_empresa is not None
        else None
    )

    with get_connection() as conn:
        claimed = _claim_records(conn, supplier, reference_date, reference_date_end)
        if not claimed:
            conn.rollback()
            raise ChargeAlreadyLaunchedError(
                f"Nenhum registro de {supplier} foi encontrado para o período "
                "selecionado. Esta cobrança provavelmente já foi lançada por "
                "outro usuário — recarregue a página para ver a base atualizada."
            )
        insert_charge_rows(conn, rows_forn, rows_emp)
        conn.commit()

    load_history.clear()
    if rows_emp is not None:
        from src.data.divida_dividida import load_dividas_divididas
        load_dividas_divididas.clear()

    from src.data.loader import load_data_from_disk
    load_data_from_disk.clear()
    _drop_supplier_from_session(supplier, reference_date, reference_date_end)

    return cod_lancamento


def _drop_supplier_from_session(
    supplier: str,
    reference_date: date | None = None,
    reference_date_end: date | None = None,
) -> None:
    """
    Remove do DataFrame ativo em session_state os registros já reivindicados no
    banco por launch_charge, para que a tela reflita a cobrança recém-lançada
    sem esperar o TTL do cache. Espelha exatamente o filtro do DELETE.
    """
    if "df" not in st.session_state:
        return

    df_atual = st.session_state["df"]
    mask_del = df_atual[COLS["supplier"]] == supplier
    if reference_date is not None:
        date_end = reference_date_end if reference_date_end is not None else reference_date
        mask_del &= (
            (df_atual[COLS["date"]].dt.date >= reference_date)
            & (df_atual[COLS["date"]].dt.date <= date_end)
        )

    st.session_state["df"] = df_atual[~mask_del].copy().reset_index(drop=True)


def update_lancamento_status(
    cod_lancamento: str,
    novo_status: str,
    data_pagamento: date | None = None,
) -> bool:
    """
    Atualiza o status de todos os itens de um lançamento e, nos status
    terminais, move-o para a tabela de destino: "Pago" → pagamentos_concluidos,
    "Devolução" → devolucoes (a oficina optou por consertar as peças em vez de
    pagar o desconto, então elas saem do fluxo de cobrança). "Pendente" apenas
    atualiza o status no lugar.

    Retorna True se o lançamento foi movido/atualizado por ESTA chamada, e False
    se o status é inválido ou se não havia nada a fazer — o código não existe, ou
    outra sessão já concluiu o mesmo lançamento. Falhas de banco propagam como
    DatabaseUnavailableError para a fronteira @page_guard da página; nunca são
    convertidas em False, senão uma indisponibilidade do Supabase ficaria
    indistinguível de "código não encontrado" e o admin tentaria de novo sem
    saber o motivo real.

    Concorrência: o UPDATE inicial é o que reivindica o lançamento. Ele trava as
    linhas, então dois admins clicando "Pago" ao mesmo tempo não duplicam o
    lançamento no destino: a segunda transação fica bloqueada, e quando a
    primeira comita (já tendo apagado as linhas) o UPDATE dela reavalia o WHERE,
    não encontra mais nada, devolve rowcount 0 e esta função retorna False.
    """
    if novo_status not in STATUS_OPTIONS:
        return False

    create_tables()

    # Uniforme para os três status: só "Pago" carrega data de pagamento. Assim o
    # INSERT ... SELECT abaixo copia as linhas já com os valores finais.
    data_pagamento_br = (
        data_pagamento.strftime("%d/%m/%Y")
        if novo_status == "Pago" and data_pagamento
        else ""
    )

    with get_connection() as conn:
        claimed = conn.execute(
            text(
                'UPDATE historico_cobrancas SET "STATUS_COBRANCA" = :s, '
                '"DATA_PAGAMENTO" = :dp WHERE "COD_LANCAMENTO" = :cod'
            ),
            {"s": novo_status, "dp": data_pagamento_br, "cod": cod_lancamento},
        ).rowcount
        if not claimed:
            return False

        destino = _TERMINAL_TABLES.get(novo_status)
        if destino:
            # INSERT ... SELECT copia direto no banco (sem trazer as linhas para
            # o Python) e omite `id`, que é PK auto-gerada no destino.
            cols_sql = ", ".join(f'"{c}"' for c in _CHARGE_COLUMNS)
            conn.execute(
                text(
                    f"INSERT INTO {destino} ({cols_sql}) SELECT {cols_sql} "
                    'FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :cod'
                ),
                {"cod": cod_lancamento},
            )
            conn.execute(
                text('DELETE FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :cod'),
                {"cod": cod_lancamento},
            )

        conn.commit()

    load_history.clear()
    if novo_status == "Pago":
        from src.data.payment_history import load_payments
        load_payments.clear()
    elif novo_status == "Devolução":
        from src.data.devolucao_history import load_devolucoes
        load_devolucoes.clear()
    return True


def _migrate_by_status(
    status_origem: str, destino: str, status_destino: str | None = None
) -> int:
    """
    Move todos os lançamentos com STATUS_COBRANCA = `status_origem` de
    historico_cobrancas para `destino`, opcionalmente reescrevendo o status para
    `status_destino`. Retorna quantas linhas foram movidas. Idempotente.

    Concorrência: roda inteira sob advisory lock porque estas migrações são
    disparadas na abertura da página de Histórico — com dois usuários entrando
    ao mesmo tempo, sem a trava as duas transações copiariam as MESMAS linhas
    para o destino antes de qualquer uma apagar a origem, duplicando pagamentos.
    """
    create_tables()

    with get_connection() as conn:
        advisory_lock(conn, f"migrate:{status_origem}")

        count = conn.execute(
            text(
                'SELECT COUNT(*) FROM historico_cobrancas '
                'WHERE "STATUS_COBRANCA" = :origem'
            ),
            {"origem": status_origem},
        ).scalar()
        if not count:
            return 0

        # A projeção do SELECT substitui status/data de pagamento quando o
        # destino exige valores diferentes dos da origem.
        select_cols = []
        for col in _CHARGE_COLUMNS:
            if col == COLS["status"] and status_destino is not None:
                select_cols.append(":status_destino")
            elif col == "DATA_PAGAMENTO" and status_destino is not None:
                select_cols.append("''")
            else:
                select_cols.append(f'"{col}"')

        insert_cols = ", ".join(f'"{c}"' for c in _CHARGE_COLUMNS)
        conn.execute(
            text(
                f"INSERT INTO {destino} ({insert_cols}) "
                f"SELECT {', '.join(select_cols)} FROM historico_cobrancas "
                'WHERE "STATUS_COBRANCA" = :origem'
            ),
            {"origem": status_origem, "status_destino": status_destino},
        )
        conn.execute(
            text('DELETE FROM historico_cobrancas WHERE "STATUS_COBRANCA" = :origem'),
            {"origem": status_origem},
        )
        conn.commit()

    load_history.clear()
    return int(count)


def migrate_paid_to_payments() -> int:
    """
    Move lançamentos com STATUS='Pago' que porventura ainda estejam em
    historico_cobrancas para pagamentos_concluidos. Idempotente.
    """
    count = _migrate_by_status("Pago", "pagamentos_concluidos")
    if count:
        from src.data.payment_history import load_payments
        load_payments.clear()
    return count


def migrate_contestado_to_devolucao() -> int:
    """
    Compatibilidade: a opção de status "Contestado" foi substituída por
    "Devolução". Move lançamentos com o status legado "Contestado" que
    porventura ainda estejam em historico_cobrancas para devolucoes,
    seguindo a mesma regra da opção atual. Idempotente.
    """
    count = _migrate_by_status("Contestado", "devolucoes", status_destino="Devolução")
    if count:
        from src.data.devolucao_history import load_devolucoes
        load_devolucoes.clear()
    return count


# ── Público: leitura ──────────────────────────────────────────────────────────

@st.cache_data(ttl=CACHE_TTL_SECONDS)
def load_history() -> pd.DataFrame | None:
    """Carrega o histórico completo de historico_cobrancas. Retorna None se vazio."""
    create_tables()
    with get_connection() as conn:
        df = pd.read_sql("SELECT * FROM historico_cobrancas", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    return df


# ── Público: geração de xlsx para download ────────────────────────────────────

def generate_history_xlsx_bytes() -> bytes | None:
    """
    Gera o xlsx formatado do histórico de cobranças em memória e retorna os bytes.
    Retorna None se não houver dados.
    """
    create_tables()
    with get_connection() as conn:
        df = pd.read_sql("SELECT * FROM historico_cobrancas", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    buf = io.BytesIO()
    _write_history_xlsx(df, buf)
    return buf.getvalue()


def generate_single_charge_xlsx_bytes(cod_lancamento: str) -> bytes | None:
    """
    Gera o xlsx formatado de um único lançamento (COD_LANCAMENTO) em memória
    e retorna os bytes. Retorna None se o código não existir.
    """
    create_tables()
    with get_connection() as conn:
        df = pd.read_sql(
            text('SELECT * FROM historico_cobrancas WHERE "COD_LANCAMENTO" = :cod'),
            conn,
            params={"cod": cod_lancamento},
        )
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    buf = io.BytesIO()
    _write_history_xlsx(df, buf)
    return buf.getvalue()


# ── Privado: escrita xlsx ─────────────────────────────────────────────────────

def _write_history_xlsx(df: pd.DataFrame, dest) -> None:
    """
    Grava df como xlsx formatado em dest (Path ou file-like / BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico Cobranças"

    all_cols  = list(df.columns)
    num_cols  = len(all_cols)
    last_col  = get_column_letter(num_cols)
    today_br  = date.today().strftime("%d/%m/%Y")
    n_records = len(df)

    def thin(color=_GRAY_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_border():
        thick = Side(style="medium", color=_PURPLE_MID)
        thin_ = Side(style="thin",   color=_PURPLE_MID)
        return Border(left=thin_, right=thin_, top=thick, bottom=thick)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "HISTÓRICO DE COBRANÇAS — DEFEITOS / REMONTES"
    c.font      = Font(name="Calibri", bold=True, size=15, color=_WHITE)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"Gerado em: {today_br}  ·  Total de registros: {n_records}"
    c.font      = Font(name="Calibri", size=10, color=_TEXT_LIGHT)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    header_row = _HEADER_OFFSET
    col_labels = {
        "COD_LANCAMENTO":  "Código",
        "DATA_COBRANCA":   "Data Cobrança",
        "DATA_VENCIMENTO": "Data Vencimento",
        "DATA_PAGAMENTO":  "Data Pagamento",
        "CNPJ_FORNECEDOR": "CNPJ",
        COLS["status"]:    "Status",
        COLS["order"]:     "OM",
        COLS["date"]:      "Data Produção",
        COLS["supplier"]:  "Fornecedor",
        COLS["quantity"]:  "Qtd",
        COLS["defect"]:    "Remonte / Defeito",
        COLS["real_cut"]:  "Real Cortado",
        COLS["minutes"]:   "Min. Gerados",
        COLS["value_brl"]: "Valor (R$)",
    }
    for idx, col in enumerate(all_cols, start=1):
        cell = ws.cell(row=header_row, column=idx)
        cell.value     = col_labels.get(col, col)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = header_border()
    ws.row_dimensions[header_row].height = 26

    value_col_name  = COLS["value_brl"]
    status_col_name = COLS["status"]
    total_value     = 0.0

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        fill_color = _PURPLE_LIGHT if row_idx % 2 == 0 else _WHITE

        for col_idx, col in enumerate(all_cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin()

            val = row.get(col)
            if pd.isna(val):
                val = ""

            if col == status_col_name:
                status_val = str(val) if val != "" else STATUS_DEFAULT
                colors     = _STATUS_COLORS.get(status_val, {"bg": _PURPLE_LIGHT, "fg": "1A1530"})
                cell.value     = status_val
                cell.fill      = PatternFill("solid", fgColor=colors["bg"])
                cell.font      = Font(name="Calibri", size=9, bold=True, color=colors["fg"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            if col == value_col_name:
                fval = float(val) if val != "" else 0.0
                total_value       += fval
                cell.value         = fval
                cell.number_format = 'R$ #,##0.00'
                cell.fill          = PatternFill("solid", fgColor=fill_color)
                cell.alignment     = Alignment(horizontal="right")
                cell.font          = Font(name="Calibri", size=9, color="1A1530")
                continue

            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Calibri", size=9)

            if col == "COD_LANCAMENTO":
                cell.font      = Font(name="Consolas", size=9, bold=True, color=_PURPLE_MID)
                cell.alignment = Alignment(horizontal="center")
            elif col == "DATA_VENCIMENTO":
                cell.alignment = Alignment(horizontal="center")
                _venc_dt = pd.to_datetime(val, format="%d/%m/%Y", errors="coerce") if val else None
                _status_atual = str(row.get(status_col_name, "")).strip()
                if (
                    _venc_dt is not None and not pd.isna(_venc_dt)
                    and _venc_dt.date() < date.today()
                    and _status_atual != "Pago"
                ):
                    cell.font = Font(name="Calibri", size=9, bold=True, color="C0392B")
            elif col == "DATA_PAGAMENTO":
                cell.alignment = Alignment(horizontal="center")
                _status_atual = str(row.get(status_col_name, "")).strip()
                if _status_atual == "Pago" and not val:
                    cell.value = ""
                    cell.fill  = PatternFill("solid", fgColor="FBE8C8")
                    cell.font  = Font(name="Calibri", size=9, italic=True, color="9A6B1E")
                elif _status_atual == "Pago" and val:
                    _dias_atraso, _atrasado = payment_punctuality(val, row.get("DATA_VENCIMENTO"))
                    if _atrasado:
                        cell.font = Font(name="Calibri", size=9, bold=True, color="C0392B")
                    elif _atrasado is False:
                        cell.font = Font(name="Calibri", size=9, bold=True, color="1D9E75")
            elif col in ("DATA_COBRANCA", COLS["date"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == "CNPJ_FORNECEDOR":
                cell.font      = Font(name="Calibri", size=9, color="1D9E75", bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col in (COLS["quantity"], COLS["real_cut"], COLS["minutes"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == COLS["order"]:
                cell.font      = Font(name="Calibri", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

            cell.value = val

        ws.row_dimensions[row_idx].height = 17

    total_row   = header_row + n_records + 1
    val_col_idx = (all_cols.index(value_col_name) + 1 if value_col_name in all_cols else num_cols)
    merge_end   = get_column_letter(val_col_idx - 1)
    _RED = "C0392B"

    ws.merge_cells(f"A{total_row}:{merge_end}{total_row}")
    lc = ws[f"A{total_row}"]
    lc.value     = "TOTAL HISTÓRICO"
    lc.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    lc.fill      = PatternFill("solid", fgColor=_RED)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border    = thin(_RED)

    tc = ws.cell(row=total_row, column=val_col_idx)
    tc.value         = total_value
    tc.number_format = 'R$ #,##0.00'
    tc.font          = Font(name="Calibri", bold=True, size=11, color=_WHITE)
    tc.fill          = PatternFill("solid", fgColor=_RED)
    tc.alignment     = Alignment(horizontal="right", vertical="center")
    tc.border        = thin(_RED)
    ws.row_dimensions[total_row].height = 22

    for col_idx in range(val_col_idx + 1, num_cols + 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=_RED)
        c.border = thin(_RED)

    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = (
        "Documento gerado automaticamente pelo sistema de Controle de Qualidade. "
        "Este histórico acumula todas as cobranças confirmadas no período."
    )
    c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 28

    for idx, col in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(col, 16)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(dest)
