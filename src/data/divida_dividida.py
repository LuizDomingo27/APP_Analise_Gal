# -*- coding: utf-8 -*-
"""
Cobrança dividida entre o fornecedor (oficina) e a empresa — tabela tb_divida_dividida.

Quando o admin opta por dividir uma cobrança, o valor é repartido por um
percentual configurável: a parte do fornecedor segue o fluxo normal
(historico_cobrancas) e a parte absorvida pela empresa fica registrada aqui.
As duas metades compartilham o mesmo COD_LANCAMENTO (rastreabilidade) e são
gravadas na MESMA transação (save_split_charge) para nunca deixar meia cobrança.

Divisão proporcional (decisão de produto): Valor (R$), Minutos Gerados e
Quantidade de peças são escalados pelo mesmo fator. A quantidade é arredondada
por linha (a coluna é `integer`); por isso a soma das peças das duas metades
pode diferir em ±1 do total original. Valor e Minutos (double precision) mantêm
a divisão exata.

Camada isolada e defensiva: falhas de banco propagam como
DatabaseUnavailableError (traduzida em database.py) e são tratadas pela
fronteira @page_guard das páginas — a UI nunca recebe traceback cru.
"""

import io
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from src.config.settings import COLS
from src.data.database import DIVIDA_DIVIDIDA_DDL, create_tables, get_connection
from src.data.cobranca_history import (
    HISTORY_LABELS,
    _COL_WIDTHS,
    build_charge_rows,
    gerar_cod_lancamento,
    STATUS_DEFAULT,
)

_TABLE = "tb_divida_dividida"

# Status fixo desta tabela — a parte da empresa não tem ciclo Pago/Devolução.
STATUS_DIVIDIDA = "Dividida"

# Colunas numéricas escaladas na divisão. OM e Real Cortado são
# identificadores/referência e NÃO são escalados.
_SCALE_COLS = (COLS["value_brl"], COLS["minutes"], COLS["quantity"])

# ── Paleta (mesma família da aba Devolução) ───────────────────────────────────
_PURPLE_DARK  = "1A1530"
_PURPLE_MID   = "534AB7"
_PURPLE_LIGHT = "EDE8FF"
_WHITE        = "FFFFFF"
_GRAY_BORDER  = "C8C0F0"
_TEXT_LIGHT   = "C8C0F0"
_TEAL         = "0F86A3"
_AMBER        = "D8932E"

_HEADER_OFFSET = 5


# ── Divisão dos valores (função pura — sem I/O, fácil de testar) ──────────────

def split_records(
    df_records: pd.DataFrame, perc_empresa: float
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Divide os registros crus de uma cobrança em duas metades proporcionais.

    `perc_empresa` é a fração absorvida pela empresa, em [0, 1]. O fornecedor
    recebe o complemento (1 - perc_empresa). Escala apenas Valor (R$), Minutos
    e Quantidade; as demais colunas (data, OM, remonte, real cortado) são
    preservadas intactas.

    Retorna (df_fornecedor, df_empresa). Para Valor e Minutos, a parte do
    fornecedor é calculada como `original - parte_empresa`, garantindo que a
    soma das duas metades seja exatamente o original (sem erro de ponto
    flutuante acumulado). A Quantidade é arredondada (coluna integer).
    """
    p = float(perc_empresa)
    if not 0.0 <= p <= 1.0:
        raise ValueError("perc_empresa deve estar entre 0 e 1 (fração).")

    df_fornecedor = df_records.copy()
    df_empresa = df_records.copy()

    for col in _SCALE_COLS:
        if col not in df_records.columns:
            continue
        orig = pd.to_numeric(df_records[col], errors="coerce").fillna(0)

        if col == COLS["quantity"]:
            orig_int = orig.round().astype("int64")
            emp = (orig * p).round().astype("int64")
            forn = (orig_int - emp).clip(lower=0)
            df_empresa[col] = emp.astype("Int64")
            df_fornecedor[col] = forn.astype("Int64")
        else:
            emp = orig * p
            forn = orig - emp
            df_empresa[col] = emp
            df_fornecedor[col] = forn

    return df_fornecedor, df_empresa


# ── Garantia de schema (idempotente, sobrevive a hot-reload) ──────────────────

def _ensure_schema() -> None:
    """
    Garante tb_divida_dividida de forma idempotente. Executa a DDL própria
    (CREATE TABLE IF NOT EXISTS) a cada escrita — barato e imune ao cache de
    create_tables(), que pode ficar preso a um schema antigo após um hot-reload
    do Streamlit sem reinício do processo (mesmo motivo de defeitos_imagens.py).
    Também chama create_tables() para garantir historico_cobrancas na primeira vez.
    """
    create_tables()
    with get_connection() as conn:
        conn.execute(text(DIVIDA_DIVIDIDA_DDL))
        conn.commit()


# ── Escrita atômica das duas metades ──────────────────────────────────────────

def save_split_charge(
    df_fornecedor: pd.DataFrame,
    df_empresa: pd.DataFrame,
    cnpj: str,
    data_cobranca: date,
    data_vencimento: date,
    cod_lancamento: str | None = None,
) -> str:
    """
    Grava, na MESMA transação, a metade do fornecedor em historico_cobrancas
    (status Pendente) e a metade da empresa em tb_divida_dividida (status
    "Dividida"). Ambas com o mesmo COD_LANCAMENTO. Retorna o código usado.

    Atomicidade: um único `get_connection()` com dois `to_sql` e um `commit`.
    Se qualquer inserção falhar, nada é comitado — nunca fica meia cobrança.
    """
    _ensure_schema()

    if cod_lancamento is None:
        cod_lancamento = gerar_cod_lancamento()

    rows_forn = build_charge_rows(
        df_fornecedor, cod_lancamento, data_cobranca, data_vencimento, cnpj,
        status=STATUS_DEFAULT,
    )
    rows_emp = build_charge_rows(
        df_empresa, cod_lancamento, data_cobranca, data_vencimento, cnpj,
        status=STATUS_DIVIDIDA,
    )

    with get_connection() as conn:
        rows_forn.to_sql("historico_cobrancas", conn, if_exists="append", index=False)
        rows_emp.to_sql(_TABLE, conn, if_exists="append", index=False)
        conn.commit()

    from src.data.cobranca_history import load_history
    load_history.clear()
    load_dividas_divididas.clear()
    return cod_lancamento


# ── Leitura ───────────────────────────────────────────────────────────────────

@st.cache_data
def load_dividas_divididas() -> pd.DataFrame | None:
    """
    Carrega todas as parcelas da empresa (tb_divida_dividida). Retorna None se
    vazio. Cacheada: o SELECT * só vai ao banco no cache-miss (a tabela é
    pequena — uma linha por item de cobrança dividida — então o full scan é
    barato e segue o mesmo padrão das abas Histórico/Devolução).
    """
    _ensure_schema()
    with get_connection() as conn:
        df = pd.read_sql(f"SELECT * FROM {_TABLE}", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    return df


# ── Geração de xlsx para download ─────────────────────────────────────────────

def generate_divida_dividida_xlsx_bytes() -> bytes | None:
    """Gera o xlsx executivo em memória e retorna os bytes. None se vazio."""
    _ensure_schema()
    with get_connection() as conn:
        df = pd.read_sql(f"SELECT * FROM {_TABLE}", conn)
    df = df.drop(columns=["id"], errors="ignore")
    if df.empty:
        return None
    df["DATA_PAGAMENTO"] = df["DATA_PAGAMENTO"].fillna("").astype(str).replace(
        {"nan": "", "None": "", "NaT": ""}
    )
    buf = io.BytesIO()
    _write_divida_dividida_xlsx(df, buf)
    return buf.getvalue()


# ── Privado — escrita xlsx (adaptado de devolucao_history._write_devolucoes_xlsx) ─

def _write_divida_dividida_xlsx(df: pd.DataFrame, dest) -> None:
    """Grava df como xlsx executivo em dest (Path ou file-like / BytesIO)."""
    df = df.reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cobrança Dividida"

    all_cols  = list(df.columns)
    num_cols  = len(all_cols)
    last_col  = get_column_letter(num_cols)
    today_br  = date.today().strftime("%d/%m/%Y")
    n_records = len(df)

    value_col_name  = COLS["value_brl"]
    status_col_name = COLS["status"]
    cod_col_name    = "COD_LANCAMENTO"

    total_value    = float(pd.to_numeric(df[value_col_name], errors="coerce").fillna(0).sum()) if value_col_name in df.columns else 0.0
    total_pieces   = int(pd.to_numeric(df[COLS["quantity"]], errors="coerce").fillna(0).sum()) if COLS["quantity"] in df.columns else 0
    n_lancamentos  = df[cod_col_name].nunique() if cod_col_name in df.columns else n_records
    n_fornecedores = df[COLS["supplier"]].nunique() if COLS["supplier"] in df.columns else 0
    n_orders       = df[COLS["order"]].nunique() if COLS["order"] in df.columns else 0

    def thin(color=_GRAY_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_border():
        thick = Side(style="medium", color=_PURPLE_MID)
        thin_ = Side(style="thin",   color=_PURPLE_MID)
        return Border(left=thin_, right=thin_, top=thick, bottom=thick)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "COBRANÇA DIVIDIDA — PARTE ABSORVIDA PELA EMPRESA"
    c.font      = Font(name="Calibri", bold=True, size=15, color=_WHITE)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"Gerado em: {today_br}  ·  Lançamentos divididos: {n_lancamentos}  ·  Itens: {n_records}"
    c.font      = Font(name="Calibri", size=10, color=_TEXT_LIGHT)
    c.fill      = PatternFill("solid", fgColor=_PURPLE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    kpis = [
        ("VALOR ABSORVIDO PELA EMPRESA", f"R$ {total_value:,.2f}", _TEAL),
        ("LANÇAMENTOS DIVIDIDOS", str(n_lancamentos), _PURPLE_MID),
        ("FORNECEDORES", str(n_fornecedores), _PURPLE_MID),
        ("PEÇAS", f"{total_pieces:,}", _AMBER),
        ("ORDENS (OM)", str(n_orders), _AMBER),
    ]
    n_kpi      = len(kpis)
    base_width = num_cols // n_kpi
    extra      = num_cols % n_kpi
    col_cursor = 1
    for i, (label, value, color) in enumerate(kpis):
        span       = max(base_width + (1 if i < extra else 0), 1)
        start_col  = col_cursor
        end_col    = min(col_cursor + span - 1, num_cols)
        col_cursor = end_col + 1

        start_letter = get_column_letter(start_col)
        end_letter   = get_column_letter(max(end_col, start_col))
        ws.merge_cells(f"{start_letter}3:{end_letter}3")
        cell           = ws[f"{start_letter}3"]
        cell.value     = f"{label}:  {value}"
        cell.font      = Font(name="Calibri", bold=True, size=10, color=color)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_LIGHT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 6

    header_row = _HEADER_OFFSET
    col_labels = dict(HISTORY_LABELS)
    col_labels["COD_LANCAMENTO"] = "Código da Divisão"

    for idx, col in enumerate(all_cols, start=1):
        cell           = ws.cell(row=header_row, column=idx)
        cell.value     = col_labels.get(col, col)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
        cell.fill      = PatternFill("solid", fgColor=_PURPLE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = header_border()
    ws.row_dimensions[header_row].height = 26

    last_cod    = None
    band_toggle = False

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        cur_cod = row.get(cod_col_name)
        if cur_cod != last_cod:
            band_toggle = not band_toggle
            last_cod = cur_cod
        fill_color = _PURPLE_LIGHT if band_toggle else _WHITE

        for col_idx, col in enumerate(all_cols, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin()

            val = row.get(col)
            if pd.isna(val):
                val = ""

            if col == status_col_name:
                cell.value     = STATUS_DIVIDIDA
                cell.fill      = PatternFill("solid", fgColor=_TEAL)
                cell.font      = Font(name="Calibri", size=9, bold=True, color=_WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            if col == value_col_name:
                fval               = float(val) if val != "" else 0.0
                cell.value         = fval
                cell.number_format = 'R$ #,##0.00'
                cell.fill          = PatternFill("solid", fgColor=fill_color)
                cell.alignment     = Alignment(horizontal="right")
                cell.font          = Font(name="Calibri", size=9, color="1A1530")
                continue

            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Calibri", size=9)

            if col == cod_col_name:
                cell.font      = Font(name="Consolas", size=9, bold=True, color=_PURPLE_MID)
                cell.alignment = Alignment(horizontal="center")
            elif col in ("DATA_COBRANCA", "DATA_VENCIMENTO", "DATA_PAGAMENTO", COLS["date"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == "CNPJ_FORNECEDOR":
                cell.font      = Font(name="Calibri", size=9, color=_TEAL, bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col in (COLS["quantity"], COLS["real_cut"], COLS["minutes"]):
                cell.alignment = Alignment(horizontal="center")
            elif col == COLS["order"]:
                cell.font      = Font(name="Calibri", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

            cell.value = val

        ws.row_dimensions[row_idx].height = 17

    total_row   = header_row + n_records + 1
    val_col_idx = (all_cols.index(value_col_name) + 1 if value_col_name in all_cols else num_cols)
    merge_end   = get_column_letter(max(val_col_idx - 1, 1))

    ws.merge_cells(f"A{total_row}:{merge_end}{total_row}")
    lc           = ws[f"A{total_row}"]
    lc.value     = "TOTAL ABSORVIDO PELA EMPRESA"
    lc.font      = Font(name="Calibri", bold=True, size=10, color=_WHITE)
    lc.fill      = PatternFill("solid", fgColor=_TEAL)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    lc.border    = thin(_TEAL)

    tc               = ws.cell(row=total_row, column=val_col_idx)
    tc.value         = total_value
    tc.number_format = 'R$ #,##0.00'
    tc.font          = Font(name="Calibri", bold=True, size=11, color=_WHITE)
    tc.fill          = PatternFill("solid", fgColor=_TEAL)
    tc.alignment     = Alignment(horizontal="right", vertical="center")
    tc.border        = thin(_TEAL)
    ws.row_dimensions[total_row].height = 22

    for col_idx in range(val_col_idx + 1, num_cols + 1):
        c        = ws.cell(row=total_row, column=col_idx)
        c.fill   = PatternFill("solid", fgColor=_TEAL)
        c.border = thin(_TEAL)

    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    c           = ws[f"A{footer_row}"]
    c.value     = (
        "Documento gerado automaticamente pelo sistema de Controle de Qualidade. "
        "Cada Código da Divisão identifica a parte de uma cobrança absorvida pela "
        "empresa; a outra parte foi cobrada do fornecedor no Histórico de Cobranças."
    )
    c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 28

    for idx, col in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(col, 16)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(dest)
